"""
AEGIS Proposal Compare API Routes

Endpoints:
- POST /api/proposal-compare/upload          — Upload proposal files, extract data
- POST /api/proposal-compare/compare         — Compare extracted proposals
- POST /api/proposal-compare/export          — Export comparison as XLSX
- POST /api/proposal-compare/export-html     — Export comparison as interactive HTML

Project Management:
- GET  /api/proposal-compare/projects        — List all projects
- POST /api/proposal-compare/projects        — Create a new project
- GET  /api/proposal-compare/projects/<id>   — Get project details
- PUT  /api/proposal-compare/projects/<id>   — Update project
- DELETE /api/proposal-compare/projects/<id> — Delete project
- GET  /api/proposal-compare/projects/<id>/proposals — List proposals in project
- POST /api/proposal-compare/projects/<id>/proposals — Add proposal to project
- GET  /api/proposal-compare/proposals/<id>          — Get full proposal data
- PUT  /api/proposal-compare/proposals/<id>          — Update proposal data (edit persistence)
- DELETE /api/proposal-compare/proposals/<id>        — Remove proposal from project
- POST /api/proposal-compare/proposals/<id>/move     — Move proposal to different project
- POST /api/proposal-compare/projects/<id>/compare   — Compare proposals in a project
- GET  /api/proposal-compare/projects/<id>/comparisons — List comparisons for project

History:
- GET  /api/proposal-compare/history         — List recent comparisons
- GET  /api/proposal-compare/history/<id>    — Get comparison detail
- DELETE /api/proposal-compare/history/<id>  — Delete comparison

Metrics:
- GET  /api/proposal-compare/metrics         — Aggregated metrics for M&A dashboard
"""

import os
import json
import logging
import tempfile
import traceback
from flask import Blueprint, request, jsonify, send_file, current_app

logger = logging.getLogger(__name__)

pc_blueprint = Blueprint('proposal_compare', __name__)


# ──────────────────────────────────────────
# Init DB on first request
# ──────────────────────────────────────────

_db_initialized = False

@pc_blueprint.before_app_request
def _ensure_db():
    global _db_initialized
    if not _db_initialized:
        try:
            from .projects import init_db
            init_db()
            _db_initialized = True
        except Exception as e:
            logger.error(f"Failed to init proposal projects DB: {e}")
            _db_initialized = True  # Don't retry every request


# ──────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────

def _cleanup_old_proposal_files(upload_dir, max_age_seconds=3600):
    """Remove proposal files older than max_age_seconds to prevent disk bloat."""
    import time
    now = time.time()
    try:
        for fname in os.listdir(upload_dir):
            fpath = os.path.join(upload_dir, fname)
            if os.path.isfile(fpath):
                age = now - os.path.getmtime(fpath)
                if age > max_age_seconds:
                    try:
                        os.unlink(fpath)
                    except Exception:
                        pass
    except Exception:
        pass


# ──────────────────────────────────────────
# Upload & Extract
# ──────────────────────────────────────────

@pc_blueprint.route('/api/proposal-compare/upload', methods=['POST'])
def upload_proposals():
    """Upload one or more proposal files and extract financial data.

    Expects multipart/form-data with 'files[]' field.
    Optional: 'project_id' form field to auto-add to a project.
    Returns extracted data for each file.
    """
    try:
        from .parser import parse_proposal, SUPPORTED_EXTENSIONS

        if 'files[]' not in request.files:
            return jsonify({
                'success': False,
                'error': {'message': 'No files provided. Use "files[]" field.'}
            }), 400

        files = request.files.getlist('files[]')
        if not files:
            return jsonify({
                'success': False,
                'error': {'message': 'No files provided'}
            }), 400

        if len(files) > 10:
            return jsonify({
                'success': False,
                'error': {'message': 'Maximum 10 files allowed per upload'}
            }), 400

        project_id = request.form.get('project_id')

        results = []

        # Store uploaded files in a persistent session directory so the doc viewer
        # can serve them via /api/proposal-compare/file/<idx> (same pattern as
        # /api/scan-history/document-file for the main review tool)
        upload_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'temp', 'proposals'
        )
        os.makedirs(upload_dir, exist_ok=True)

        # Clean up old proposal files (>1 hour old) to prevent disk bloat
        _cleanup_old_proposal_files(upload_dir, max_age_seconds=3600)

        # Track saved file paths for this session
        session_files = []

        for f in files:
            if not f.filename:
                continue

            ext = os.path.splitext(f.filename)[1].lower()
            if ext not in SUPPORTED_EXTENSIONS:
                results.append({
                    'filename': f.filename,
                    'success': False,
                    'error': f'Unsupported file type: {ext}. Supported: {", ".join(SUPPORTED_EXTENSIONS)}',
                })
                continue

            # Save file to persistent temp directory with unique prefix
            import time
            safe_name = f'{int(time.time())}_{f.filename}'
            temp_path = os.path.join(upload_dir, safe_name)
            try:
                f.save(temp_path)
            except Exception as save_err:
                results.append({
                    'filename': f.filename,
                    'success': False,
                    'error': f'Failed to save file: {save_err}',
                })
                continue

            # Parse
            try:
                proposal_data = parse_proposal(temp_path)
                data_dict = proposal_data.to_dict()
                # Store the server-side file path so doc viewer can serve it
                data_dict['_server_file'] = safe_name

                # Auto-add to project if specified
                db_id = None
                if project_id:
                    try:
                        from .projects import add_proposal_to_project
                        result_row = add_proposal_to_project(int(project_id), data_dict)
                        db_id = result_row.get('id')
                    except Exception as db_err:
                        logger.warning(f'Failed to add proposal to project: {db_err}')

                session_files.append(temp_path)
                results.append({
                    'filename': f.filename,
                    'success': True,
                    'data': data_dict,
                    'db_id': db_id,
                })
            except Exception as parse_err:
                logger.error(f'Proposal parse error for {f.filename}: {parse_err}', exc_info=True)
                results.append({
                    'filename': f.filename,
                    'success': False,
                    'error': f'Parse error: {str(parse_err)}',
                })
                # Clean up file on parse failure
                try:
                    os.unlink(temp_path)
                except Exception:
                    pass

        successful = sum(1 for r in results if r.get('success'))

        return jsonify({
            'success': True,
            'data': {
                'results': results,
                'total_uploaded': len(files),
                'successful': successful,
                'failed': len(files) - successful,
            }
        })

    except Exception as e:
        logger.error(f'Proposal upload error: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': {'message': f'Upload error: {str(e)}', 'traceback': traceback.format_exc()}
        }), 500


# ──────────────────────────────────────────
# Serve Uploaded File (for doc viewer)
# ──────────────────────────────────────────

@pc_blueprint.route('/api/proposal-compare/file/<path:filename>', methods=['GET'])
def serve_proposal_file(filename):
    """Serve an uploaded proposal file for the document viewer.

    Same pattern as /api/scan-history/document-file — serves the original
    file from disk so PDF.js can render it via a server URL (not blob URL).
    """
    try:
        # Sanitize filename — only allow the basename, no path traversal
        safe_name = os.path.basename(filename)
        if not safe_name or '..' in safe_name:
            return jsonify({'success': False, 'error': 'Invalid filename'}), 400

        upload_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'temp', 'proposals'
        )
        filepath = os.path.join(upload_dir, safe_name)

        if not os.path.isfile(filepath):
            return jsonify({'success': False, 'error': 'File not found'}), 404

        return send_file(filepath, as_attachment=False)

    except Exception as e:
        logger.error(f'Serve proposal file error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ──────────────────────────────────────────
# Compare
# ──────────────────────────────────────────

@pc_blueprint.route('/api/proposal-compare/compare', methods=['POST'])
def compare_proposals_endpoint():
    """Compare extracted proposal data.

    Expects JSON body with 'proposals' array of ProposalData dicts
    (as returned by the upload endpoint).
    Optional: 'project_id' to save comparison result.
    """
    try:
        from .parser import ProposalData, LineItem, ExtractedTable
        from .analyzer import compare_proposals

        data = request.get_json()
        if not data or 'proposals' not in data:
            return jsonify({
                'success': False,
                'error': {'message': 'Missing "proposals" array in request body'}
            }), 400

        raw_proposals = data['proposals']
        if len(raw_proposals) < 2:
            return jsonify({
                'success': False,
                'error': {'message': 'Need at least 2 proposals to compare'}
            }), 400

        if len(raw_proposals) > 10:
            return jsonify({
                'success': False,
                'error': {'message': 'Maximum 10 proposals allowed'}
            }), 400

        # Reconstruct ProposalData objects from dicts
        proposals = []
        for rp in raw_proposals:
            p = ProposalData(
                filename=rp.get('filename', ''),
                filepath=rp.get('filepath', ''),
                file_type=rp.get('file_type', ''),
                company_name=rp.get('company_name', ''),
                proposal_title=rp.get('proposal_title', ''),
                date=rp.get('date', ''),
                total_amount=rp.get('total_amount'),
                total_raw=rp.get('total_raw', ''),
                currency=rp.get('currency', 'USD'),
                page_count=rp.get('page_count', 0),
                extraction_notes=rp.get('extraction_notes', []),
                contract_term=rp.get('contract_term', ''),
                extraction_text=rp.get('extraction_text', ''),
            )

            # Reconstruct tables
            for td in rp.get('tables', []):
                t = ExtractedTable(
                    headers=td.get('headers', []),
                    rows=td.get('rows', []),
                    source=td.get('source', ''),
                    table_index=td.get('table_index', 0),
                    has_financial_data=td.get('has_financial_data', False),
                    total_row_index=td.get('total_row_index'),
                )
                p.tables.append(t)

            # Reconstruct line items
            for lid in rp.get('line_items', []):
                li = LineItem(
                    description=lid.get('description', ''),
                    amount=lid.get('amount'),
                    amount_raw=lid.get('amount_raw', ''),
                    quantity=lid.get('quantity'),
                    unit_price=lid.get('unit_price'),
                    unit=lid.get('unit', ''),
                    category=lid.get('category', ''),
                    row_index=lid.get('row_index', 0),
                    table_index=lid.get('table_index', 0),
                    source_sheet=lid.get('source_sheet', ''),
                    confidence=lid.get('confidence', 1.0),
                )
                p.line_items.append(li)

            proposals.append(p)

        # Run comparison
        result = compare_proposals(proposals)
        result_dict = result.to_dict()

        # Auto-save all comparisons to history
        comparison_id = None
        try:
            from .projects import save_comparison
            project_id = data.get('project_id')
            proposal_ids = data.get('proposal_db_ids', [])
            term_label = data.get('term_label', '')
            # Store raw proposals for history re-editing
            proposals_input = data.get('proposals', [])
            # Build notes from term_label for multi-term tracking
            notes = f'Term: {term_label}' if term_label else ''
            comparison_id = save_comparison(
                int(project_id) if project_id else 0,
                proposal_ids,
                result_dict,
                notes=notes,
                proposals_json=proposals_input,
            )
        except Exception as db_err:
            logger.warning(f'Failed to save comparison: {db_err}')

        result_dict['comparison_id'] = comparison_id
        if term_label:
            result_dict['term_label'] = term_label

        return jsonify({
            'success': True,
            'data': result_dict,
        })

    except Exception as e:
        logger.error(f'Proposal compare error: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': {'message': f'Compare error: {str(e)}', 'traceback': traceback.format_exc()}
        }), 500


# ──────────────────────────────────────────
# Comparison History
# ──────────────────────────────────────────

@pc_blueprint.route('/api/proposal-compare/history', methods=['GET'])
def list_comparison_history():
    """List recent comparisons with summary metadata."""
    try:
        from .projects import list_comparisons
        limit = request.args.get('limit', 20, type=int)
        comparisons = list_comparisons(limit=limit)
        return jsonify({'success': True, 'data': comparisons})
    except Exception as e:
        logger.error(f'Failed to list comparisons: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': {'message': f'Failed to load history: {str(e)}'}
        }), 500


@pc_blueprint.route('/api/proposal-compare/history/<int:comparison_id>', methods=['GET'])
def get_comparison_detail(comparison_id):
    """Get full comparison result for re-rendering."""
    try:
        from .projects import get_comparison
        result = get_comparison(comparison_id)
        if not result:
            return jsonify({
                'success': False,
                'error': {'message': 'Comparison not found'}
            }), 404
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f'Failed to get comparison {comparison_id}: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': {'message': f'Failed to load comparison: {str(e)}'}
        }), 500


@pc_blueprint.route('/api/proposal-compare/history/<int:comparison_id>', methods=['DELETE'])
def delete_comparison_history(comparison_id):
    """Delete a saved comparison."""
    try:
        from .projects import delete_comparison
        success = delete_comparison(comparison_id)
        if not success:
            return jsonify({
                'success': False,
                'error': {'message': 'Comparison not found'}
            }), 404
        return jsonify({'success': True, 'message': 'Comparison deleted'})
    except Exception as e:
        logger.error(f'Failed to delete comparison {comparison_id}: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': {'message': f'Failed to delete comparison: {str(e)}'}
        }), 500


# ──────────────────────────────────────────
# Export
# ──────────────────────────────────────────

@pc_blueprint.route('/api/proposal-compare/export', methods=['POST'])
def export_comparison():
    """Export comparison results as XLSX."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': {'message': 'Missing comparison data'}
            }), 400

        proposals = data.get('proposals', [])
        aligned_items = data.get('aligned_items', [])
        category_summaries = data.get('category_summaries', [])
        totals = data.get('totals', {})
        red_flags = data.get('red_flags', {})
        vendor_scores = data.get('vendor_scores', {})
        executive_summary = data.get('executive_summary', {})

        # Create workbook
        wb = openpyxl.Workbook()

        # Styles
        header_fill = PatternFill(start_color='1B2838', end_color='1B2838', fill_type='solid')
        header_font = Font(color='D6A84A', bold=True, size=11)
        gold_fill = PatternFill(start_color='D6A84A', end_color='D6A84A', fill_type='solid')
        gold_font = Font(color='1B2838', bold=True, size=12)
        low_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        high_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        total_fill = PatternFill(start_color='D6A84A', end_color='D6A84A', fill_type='solid')
        total_font = Font(color='1B2838', bold=True, size=11)
        red_flag_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        warning_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='555555'),
            right=Side(style='thin', color='555555'),
            top=Side(style='thin', color='555555'),
            bottom=Side(style='thin', color='555555'),
        )
        money_fmt = '$#,##0.00'
        pct_fmt = '0.0%'

        prop_ids = [p.get('id', p.get('filename', f'Proposal {i+1}')) for i, p in enumerate(proposals)]

        # ── Sheet 1: Executive Summary ──
        ws_exec = wb.active
        ws_exec.title = 'Executive Summary'

        ws_exec.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        cell = ws_exec.cell(row=1, column=1, value='AEGIS Proposal Comparison — Executive Summary')
        cell.fill = gold_fill
        cell.font = gold_font
        cell.alignment = Alignment(horizontal='center')

        row = 3
        # Price ranking
        price_ranking = executive_summary.get('price_ranking', [])
        if price_ranking:
            ws_exec.cell(row=row, column=1, value='Price Ranking').font = Font(bold=True, size=12)
            row += 1
            for h_idx, h in enumerate(['Rank', 'Vendor', 'Total Price', 'Delta from Lowest']):
                c = ws_exec.cell(row=row, column=h_idx+1, value=h)
                c.fill = header_fill
                c.font = header_font
                c.border = border
            row += 1
            for pr in price_ranking:
                ws_exec.cell(row=row, column=1, value=pr['rank']).border = border
                ws_exec.cell(row=row, column=2, value=pr['vendor']).border = border
                c = ws_exec.cell(row=row, column=3, value=pr['total'])
                c.number_format = money_fmt
                c.border = border
                if pr['rank'] == 1:
                    c.fill = low_fill
                delta = ws_exec.cell(row=row, column=4, value=f"+{pr['delta_pct']}%" if pr['delta_pct'] > 0 else '—')
                delta.border = border
                row += 1

        # Vendor scores
        row += 1
        if vendor_scores:
            ws_exec.cell(row=row, column=1, value='Vendor Scores').font = Font(bold=True, size=12)
            row += 1
            for h_idx, h in enumerate(['Vendor', 'Overall', 'Grade', 'Price', 'Complete', 'Risk', 'Red Flags']):
                c = ws_exec.cell(row=row, column=h_idx+1, value=h)
                c.fill = header_fill
                c.font = header_font
                c.border = border
            row += 1
            for pid in prop_ids:
                vs = vendor_scores.get(pid, {})
                ws_exec.cell(row=row, column=1, value=pid).border = border
                ws_exec.cell(row=row, column=2, value=vs.get('overall', 0)).border = border
                ws_exec.cell(row=row, column=3, value=vs.get('grade', '—')).border = border
                ws_exec.cell(row=row, column=4, value=vs.get('price_score', 0)).border = border
                ws_exec.cell(row=row, column=5, value=vs.get('completeness_score', 0)).border = border
                ws_exec.cell(row=row, column=6, value=vs.get('risk_score', 0)).border = border
                ws_exec.cell(row=row, column=7, value=vs.get('red_flag_count', 0)).border = border
                row += 1

        # Negotiation opportunities
        row += 1
        negot = executive_summary.get('negotiation_opportunities', [])
        if negot:
            ws_exec.cell(row=row, column=1, value='Top Negotiation Opportunities').font = Font(bold=True, size=12)
            row += 1
            for h_idx, h in enumerate(['Vendor', 'Line Item', 'Current', 'Average', 'Potential Savings']):
                c = ws_exec.cell(row=row, column=h_idx+1, value=h)
                c.fill = header_fill
                c.font = header_font
                c.border = border
            row += 1
            for n in negot[:5]:
                ws_exec.cell(row=row, column=1, value=n['vendor']).border = border
                ws_exec.cell(row=row, column=2, value=n['line_item']).border = border
                c = ws_exec.cell(row=row, column=3, value=n['current_amount'])
                c.number_format = money_fmt
                c.border = border
                c = ws_exec.cell(row=row, column=4, value=n['avg_amount'])
                c.number_format = money_fmt
                c.border = border
                c = ws_exec.cell(row=row, column=5, value=n['potential_savings'])
                c.number_format = money_fmt
                c.fill = low_fill
                c.border = border
                row += 1

        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_exec.column_dimensions[col_letter].width = 22

        # ── Sheet 2: Side-by-Side Comparison ──
        ws = wb.create_sheet('Proposal Comparison')

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(prop_ids))
        title_cell = ws.cell(row=1, column=1, value='AEGIS Proposal Comparison')
        title_cell.fill = gold_fill
        title_cell.font = gold_font
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['Line Item', 'Category'] + prop_ids + ['Variance']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = border

        row_num = 4
        for item in aligned_items:
            ws.cell(row=row_num, column=1, value=item.get('description', '')).border = border
            ws.cell(row=row_num, column=2, value=item.get('category', '')).border = border

            amounts = item.get('amounts', {})
            valid_amounts = [a for a in amounts.values() if a is not None and a > 0]
            min_amount = min(valid_amounts) if valid_amounts else None
            max_amount = max(valid_amounts) if valid_amounts else None

            for p_idx, pid in enumerate(prop_ids):
                col = 3 + p_idx
                amount = amounts.get(pid)
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                if amount is not None:
                    cell.value = amount
                    cell.number_format = money_fmt
                    if len(valid_amounts) > 1:
                        if amount == min_amount:
                            cell.fill = low_fill
                        elif amount == max_amount:
                            cell.fill = high_fill
                else:
                    cell.value = '—'
                    cell.alignment = Alignment(horizontal='center')

            variance_col = 3 + len(prop_ids)
            cell = ws.cell(row=row_num, column=variance_col)
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            variance_pct = item.get('variance_pct')
            if variance_pct is not None:
                cell.value = variance_pct / 100
                cell.number_format = pct_fmt
            else:
                cell.value = '—'
                cell.alignment = Alignment(horizontal='center')
            row_num += 1

        # Grand total row
        row_num += 1
        total_cell = ws.cell(row=row_num, column=1, value='GRAND TOTAL')
        total_cell.fill = total_fill
        total_cell.font = total_font
        total_cell.border = border
        ws.cell(row=row_num, column=2, value='').fill = total_fill
        ws.cell(row=row_num, column=2).border = border

        for p_idx, pid in enumerate(prop_ids):
            col = 3 + p_idx
            cell = ws.cell(row=row_num, column=col)
            total_val = totals.get(pid)
            cell.value = total_val if total_val else '—'
            if isinstance(total_val, (int, float)):
                cell.number_format = money_fmt
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border
            cell.alignment = Alignment(horizontal='right')

        variance_col = 3 + len(prop_ids)
        cell = ws.cell(row=row_num, column=variance_col)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
        total_variance = data.get('total_variance_pct')
        if total_variance is not None:
            cell.value = total_variance / 100
            cell.number_format = pct_fmt
        else:
            cell.value = '—'

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        for p_idx in range(len(prop_ids)):
            col_letter = openpyxl.utils.get_column_letter(3 + p_idx)
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions[openpyxl.utils.get_column_letter(variance_col)].width = 12

        # Freeze panes (freeze row 3 header + column A)
        ws.freeze_panes = 'C4'
        # Auto-filter on header row
        last_col = openpyxl.utils.get_column_letter(variance_col)
        ws.auto_filter.ref = f'A3:{last_col}{row_num}'

        # ── Sheet 3: Red Flags ──
        ws3 = wb.create_sheet('Red Flags')
        rf_headers = ['Vendor', 'Severity', 'Flag Type', 'Title', 'Detail']
        for col, h in enumerate(rf_headers, 1):
            c = ws3.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.border = border
        rf_row = 2
        for pid in prop_ids:
            for flag in red_flags.get(pid, []):
                ws3.cell(row=rf_row, column=1, value=pid).border = border
                sev_cell = ws3.cell(row=rf_row, column=2, value=flag.get('severity', '').upper())
                sev_cell.border = border
                if flag.get('severity') == 'critical':
                    sev_cell.fill = red_flag_fill
                elif flag.get('severity') == 'warning':
                    sev_cell.fill = warning_fill
                ws3.cell(row=rf_row, column=3, value=flag.get('type', '')).border = border
                ws3.cell(row=rf_row, column=4, value=flag.get('title', '')).border = border
                ws3.cell(row=rf_row, column=5, value=flag.get('detail', '')).border = border
                rf_row += 1
        ws3.column_dimensions['A'].width = 25
        ws3.column_dimensions['B'].width = 12
        ws3.column_dimensions['C'].width = 20
        ws3.column_dimensions['D'].width = 30
        ws3.column_dimensions['E'].width = 60

        # ── Sheet 4: Category Summary ──
        ws4 = wb.create_sheet('Category Summary')
        cat_headers = ['Category', 'Items'] + prop_ids
        for col, h in enumerate(cat_headers, 1):
            c = ws4.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.border = border
        for row_idx, cat in enumerate(category_summaries, 2):
            ws4.cell(row=row_idx, column=1, value=cat.get('category', '')).border = border
            ws4.cell(row=row_idx, column=2, value=cat.get('item_count', 0)).border = border
            cat_totals = cat.get('totals', {})
            for p_idx, pid in enumerate(prop_ids):
                c = ws4.cell(row=row_idx, column=3 + p_idx)
                val = cat_totals.get(pid, 0)
                c.value = val if val else '—'
                if isinstance(val, (int, float)) and val > 0:
                    c.number_format = money_fmt
                c.border = border

        ws4.column_dimensions['A'].width = 20
        ws4.column_dimensions['B'].width = 10
        for p_idx in range(len(prop_ids)):
            col_letter = openpyxl.utils.get_column_letter(3 + p_idx)
            ws4.column_dimensions[col_letter].width = 20

        # ── Sheet 5: Proposal Details ──
        ws5 = wb.create_sheet('Proposal Details')
        detail_headers = ['Field'] + prop_ids
        for col, h in enumerate(detail_headers, 1):
            c = ws5.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.border = border

        detail_fields = [
            ('Company', 'company_name'),
            ('Title', 'proposal_title'),
            ('Date', 'date'),
            ('File', 'filename'),
            ('Type', 'file_type'),
            ('Tables Found', 'table_count'),
            ('Line Items', 'line_item_count'),
            ('Grand Total', 'total_raw'),
        ]

        for row_idx, (label, key) in enumerate(detail_fields, 2):
            ws5.cell(row=row_idx, column=1, value=label).border = border
            for p_idx, p in enumerate(proposals):
                c = ws5.cell(row=row_idx, column=2 + p_idx, value=str(p.get(key, '')))
                c.border = border

        ws5.column_dimensions['A'].width = 15
        for p_idx in range(len(prop_ids)):
            col_letter = openpyxl.utils.get_column_letter(2 + p_idx)
            ws5.column_dimensions[col_letter].width = 30

        # ── Sheet 6: Heatmap ──
        heatmap_data = data.get('heatmap', {})
        heatmap_categories = heatmap_data.get('categories', [])
        heatmap_vendors = heatmap_data.get('vendors', [])
        heatmap_grid = heatmap_data.get('grid', {})
        if heatmap_categories and heatmap_vendors:
            ws6 = wb.create_sheet('Heatmap')
            # Header row
            ws6.cell(row=1, column=1, value='Category / Vendor').fill = header_fill
            ws6.cell(row=1, column=1).font = header_font
            ws6.cell(row=1, column=1).border = border
            for v_idx, vendor in enumerate(heatmap_vendors):
                c = ws6.cell(row=1, column=2 + v_idx, value=vendor)
                c.fill = header_fill
                c.font = header_font
                c.border = border
            # Data rows
            heatmap_low = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            heatmap_mid = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            heatmap_high = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            for c_idx, cat in enumerate(heatmap_categories):
                ws6.cell(row=2 + c_idx, column=1, value=cat).border = border
                ws6.cell(row=2 + c_idx, column=1).font = Font(bold=True)
                cat_grid = heatmap_grid.get(cat, {})
                for v_idx, vendor in enumerate(heatmap_vendors):
                    cell = ws6.cell(row=2 + c_idx, column=2 + v_idx)
                    cell.border = border
                    cell_data = cat_grid.get(vendor, {})
                    amount = cell_data.get('amount', 0) if isinstance(cell_data, dict) else cell_data
                    deviation = cell_data.get('deviation', 0) if isinstance(cell_data, dict) else 0
                    if amount and isinstance(amount, (int, float)):
                        cell.value = amount
                        cell.number_format = money_fmt
                        # Color by deviation from average
                        if deviation <= -10:
                            cell.fill = heatmap_low  # Below average (green = good)
                        elif deviation >= 10:
                            cell.fill = heatmap_high  # Above average (red = expensive)
                        elif deviation != 0:
                            cell.fill = heatmap_mid  # Near average
                    else:
                        cell.value = '—'
                        cell.alignment = Alignment(horizontal='center')
            ws6.column_dimensions['A'].width = 20
            for v_idx in range(len(heatmap_vendors)):
                ws6.column_dimensions[openpyxl.utils.get_column_letter(2 + v_idx)].width = 20

        # ── Sheet 7: Rate Analysis ──
        rate_analysis = data.get('rate_analysis', {})
        indirect_rates = data.get('indirect_rates', {})
        if indirect_rates:
            ws7 = wb.create_sheet('Rate Analysis')
            ws7.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(prop_ids))
            c = ws7.cell(row=1, column=1, value='Indirect Rate Analysis')
            c.fill = gold_fill
            c.font = gold_font
            c.alignment = Alignment(horizontal='center')

            rate_types = ['overhead_rate', 'ga_rate', 'fringe_rate', 'fee_rate']
            rate_labels = ['Overhead Rate', 'G&A Rate', 'Fringe Rate', 'Fee / Profit Rate']
            typical_ranges = {
                'overhead_rate': (0.40, 1.20),
                'ga_rate': (0.10, 0.25),
                'fringe_rate': (0.25, 0.45),
                'fee_rate': (0.08, 0.15),
            }

            # Headers
            rate_headers = ['Rate Type', 'Typical Range'] + prop_ids
            for col, h in enumerate(rate_headers, 1):
                c = ws7.cell(row=3, column=col, value=h)
                c.fill = header_fill
                c.font = header_font
                c.border = border

            for r_idx, (rtype, rlabel) in enumerate(zip(rate_types, rate_labels)):
                row = 4 + r_idx
                ws7.cell(row=row, column=1, value=rlabel).border = border
                typical = typical_ranges.get(rtype, (0, 1))
                ws7.cell(row=row, column=2, value=f'{typical[0]*100:.0f}% - {typical[1]*100:.0f}%').border = border

                for p_idx, pid in enumerate(prop_ids):
                    cell = ws7.cell(row=row, column=3 + p_idx)
                    cell.border = border
                    vendor_rates = indirect_rates.get(pid, {})
                    rate_val = vendor_rates.get(rtype)
                    if rate_val is not None:
                        cell.value = rate_val
                        cell.number_format = pct_fmt
                        # Flag outside typical range
                        if rate_val < typical[0] or rate_val > typical[1]:
                            cell.fill = warning_fill
                    else:
                        cell.value = '—'
                        cell.alignment = Alignment(horizontal='center')

            ws7.column_dimensions['A'].width = 22
            ws7.column_dimensions['B'].width = 18
            for p_idx in range(len(prop_ids)):
                ws7.column_dimensions[openpyxl.utils.get_column_letter(3 + p_idx)].width = 20

        # ── Sheet 8: Raw Line Items ──
        ws8 = wb.create_sheet('Raw Line Items')
        raw_headers = ['Vendor', 'Description', 'Category', 'Amount', 'Quantity', 'Unit Price', 'Source Sheet']
        for col, h in enumerate(raw_headers, 1):
            c = ws8.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.border = border
        raw_row = 2
        for p_idx, p in enumerate(proposals):
            pid = prop_ids[p_idx]
            for li in p.get('line_items', []):
                ws8.cell(row=raw_row, column=1, value=pid).border = border
                ws8.cell(row=raw_row, column=2, value=li.get('description', '')).border = border
                ws8.cell(row=raw_row, column=3, value=li.get('category', 'Other')).border = border
                amt_cell = ws8.cell(row=raw_row, column=4)
                amt_val = li.get('amount')
                if amt_val is not None and isinstance(amt_val, (int, float)):
                    amt_cell.value = amt_val
                    amt_cell.number_format = money_fmt
                else:
                    amt_cell.value = '—'
                amt_cell.border = border
                ws8.cell(row=raw_row, column=5, value=li.get('quantity', '')).border = border
                up_cell = ws8.cell(row=raw_row, column=6)
                up_val = li.get('unit_price')
                if up_val is not None and isinstance(up_val, (int, float)):
                    up_cell.value = up_val
                    up_cell.number_format = money_fmt
                else:
                    up_cell.value = '—'
                up_cell.border = border
                ws8.cell(row=raw_row, column=7, value=li.get('source_sheet', '')).border = border
                raw_row += 1

        ws8.column_dimensions['A'].width = 25
        ws8.column_dimensions['B'].width = 50
        ws8.column_dimensions['C'].width = 15
        ws8.column_dimensions['D'].width = 16
        ws8.column_dimensions['E'].width = 10
        ws8.column_dimensions['F'].width = 16
        ws8.column_dimensions['G'].width = 20
        ws8.auto_filter.ref = f'A1:G{raw_row - 1}'
        ws8.freeze_panes = 'A2'

        # Add grade coloring to vendor scores section in Exec Summary
        grade_fills = {
            'A': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'B': PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),
            'C': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'D': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
            'F': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        }
        # Walk exec summary vendor scores to colorize grade cells
        for r in range(1, ws_exec.max_row + 1):
            grade_val = ws_exec.cell(row=r, column=3).value
            if grade_val in grade_fills:
                ws_exec.cell(row=r, column=3).fill = grade_fills[grade_val]

        # Save to temp file
        temp_path = os.path.join(tempfile.gettempdir(), 'aegis_proposal_comparison.xlsx')
        wb.save(temp_path)

        return send_file(
            temp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='AEGIS_Proposal_Comparison.xlsx'
        )

    except Exception as e:
        logger.error(f'Proposal export error: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': {'message': f'Export error: {str(e)}', 'traceback': traceback.format_exc()}
        }), 500


# ──────────────────────────────────────────
# HTML Export Endpoint
# ──────────────────────────────────────────

@pc_blueprint.route('/api/proposal-compare/export-html', methods=['POST'])
def export_html():
    """Export comparison results as interactive standalone HTML."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': {'message': 'Missing comparison data'}
            }), 400

        try:
            from proposal_compare_export import generate_proposal_compare_html
        except ImportError:
            return jsonify({
                'success': False,
                'error': {'message': 'HTML export module not available (proposal_compare_export.py)'}
            }), 500

        html_content = generate_proposal_compare_html(data)

        # Create temp file
        project_name = data.get('metadata', {}).get('project_name', 'Comparison')
        safe_name = ''.join(c if c.isalnum() or c in ' _-' else '' for c in project_name).strip()
        from datetime import datetime as dt
        timestamp = dt.now().strftime('%Y%m%d')
        filename = f'AEGIS_Proposal_Comparison_{safe_name}_{timestamp}.html'

        temp_path = os.path.join(tempfile.gettempdir(), filename)
        with open(temp_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return send_file(
            temp_path,
            mimetype='text/html',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        logger.error(f'HTML export error: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': {'message': f'HTML export error: {str(e)}', 'traceback': traceback.format_exc()}
        }), 500


# ──────────────────────────────────────────
# Project Management Endpoints
# ──────────────────────────────────────────

@pc_blueprint.route('/api/proposal-compare/projects', methods=['GET'])
def list_projects():
    """List all proposal comparison projects."""
    try:
        from .projects import list_projects as _list
        status = request.args.get('status', 'active')
        projects = _list(status)
        return jsonify({'success': True, 'data': projects})
    except Exception as e:
        logger.error(f'List projects error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/projects', methods=['POST'])
def create_project():
    """Create a new proposal comparison project."""
    try:
        from .projects import create_project as _create
        data = request.get_json()
        if not data or not data.get('name'):
            return jsonify({
                'success': False,
                'error': {'message': 'Project name is required'}
            }), 400
        project = _create(data['name'], data.get('description', ''))
        return jsonify({'success': True, 'data': project})
    except Exception as e:
        logger.error(f'Create project error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/projects/<int:project_id>', methods=['GET'])
def get_project(project_id):
    """Get a single project's details."""
    try:
        from .projects import get_project as _get
        project = _get(project_id)
        if not project:
            return jsonify({'success': False, 'error': {'message': 'Project not found'}}), 404
        return jsonify({'success': True, 'data': project})
    except Exception as e:
        logger.error(f'Get project error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/projects/<int:project_id>', methods=['PUT'])
def update_project(project_id):
    """Update a project's metadata."""
    try:
        from .projects import update_project as _update
        data = request.get_json()
        project = _update(
            project_id,
            name=data.get('name'),
            description=data.get('description'),
            status=data.get('status'),
        )
        if not project:
            return jsonify({'success': False, 'error': {'message': 'Project not found'}}), 404
        return jsonify({'success': True, 'data': project})
    except Exception as e:
        logger.error(f'Update project error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    """Delete a project and all its proposals."""
    try:
        from .projects import delete_project as _delete
        _delete(project_id)
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f'Delete project error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/projects/<int:project_id>/proposals', methods=['GET'])
def get_project_proposals(project_id):
    """List all proposals in a project."""
    try:
        from .projects import get_project_proposals as _get_proposals
        proposals = _get_proposals(project_id)
        return jsonify({'success': True, 'data': proposals})
    except Exception as e:
        logger.error(f'Get proposals error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/projects/<int:project_id>/proposals', methods=['POST'])
def add_proposal_to_project(project_id):
    """Add an already-extracted proposal to a project (from JSON body)."""
    try:
        from .projects import add_proposal_to_project as _add
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': {'message': 'Proposal data required'}
            }), 400
        result = _add(project_id, data)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f'Add proposal error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/proposals/<int:proposal_id>', methods=['DELETE'])
def remove_proposal(proposal_id):
    """Remove a proposal from its project."""
    try:
        from .projects import remove_proposal_from_project as _remove
        _remove(proposal_id)
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f'Remove proposal error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/proposals/<int:proposal_id>', methods=['GET'])
def get_proposal_data(proposal_id):
    """Get the full extracted data for a single proposal (for editing from dashboard)."""
    try:
        from .projects import get_proposal_full_data
        data = get_proposal_full_data(proposal_id)
        if not data:
            return jsonify({
                'success': False,
                'error': {'message': 'Proposal not found'}
            }), 404
        # Include the db_id so frontend can track it
        data['_db_id'] = proposal_id
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        logger.error(f'Get proposal data error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/proposals/<int:proposal_id>', methods=['PUT'])
def update_proposal(proposal_id):
    """Update a proposal's extracted data after user edits.

    Persists changes from the review phase editor (company name, line items,
    totals, etc.) back to the database so they survive modal close/reopen.
    """
    try:
        from .projects import update_proposal_data
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': {'message': 'Updated proposal data required'}
            }), 400
        result = update_proposal_data(proposal_id, data)
        if not result:
            return jsonify({
                'success': False,
                'error': {'message': 'Proposal not found'}
            }), 404
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.error(f'Update proposal error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/proposals/<int:proposal_id>/move', methods=['POST'])
def move_proposal_endpoint(proposal_id):
    """Move a proposal to a different project.

    JSON body: { "project_id": <int> }
    """
    try:
        from .projects import move_proposal
        data = request.get_json()
        if not data or not data.get('project_id'):
            return jsonify({
                'success': False,
                'error': {'message': 'Target project_id is required'}
            }), 400
        result = move_proposal(proposal_id, int(data['project_id']))
        if not result:
            return jsonify({
                'success': False,
                'error': {'message': 'Proposal not found'}
            }), 404
        return jsonify({'success': True, 'data': result})
    except ValueError as ve:
        return jsonify({
            'success': False,
            'error': {'message': str(ve)}
        }), 404
    except Exception as e:
        logger.error(f'Move proposal error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/projects/<int:project_id>/comparisons', methods=['GET'])
def get_project_comparisons(project_id):
    """List comparisons for a specific project."""
    try:
        from .projects import list_comparisons_for_project
        limit = request.args.get('limit', 20, type=int)
        comparisons = list_comparisons_for_project(project_id, limit=limit)
        return jsonify({'success': True, 'data': comparisons})
    except Exception as e:
        logger.error(f'Get project comparisons error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


@pc_blueprint.route('/api/proposal-compare/projects/<int:project_id>/compare', methods=['POST'])
def compare_project_proposals(project_id):
    """Compare all proposals in a project (or selected ones).

    Optional JSON body: { "proposal_ids": [1, 2, 3] } to compare specific proposals.
    If no body, compares all proposals in the project.
    """
    try:
        from .parser import ProposalData, LineItem, ExtractedTable
        from .analyzer import compare_proposals
        from .projects import get_project_proposals, get_proposal_full_data, save_comparison

        data = request.get_json() or {}
        selected_ids = data.get('proposal_ids')

        # Get proposals from project
        all_proposals = get_project_proposals(project_id)
        if selected_ids:
            all_proposals = [p for p in all_proposals if p['id'] in selected_ids]

        if len(all_proposals) < 2:
            return jsonify({
                'success': False,
                'error': {'message': f'Need at least 2 proposals to compare. Project has {len(all_proposals)}.'}
            }), 400

        # Load full data for each
        proposals = []
        proposal_db_ids = []
        for prop_summary in all_proposals:
            full_data = get_proposal_full_data(prop_summary['id'])
            if not full_data:
                continue

            p = ProposalData(
                filename=full_data.get('filename', ''),
                filepath=full_data.get('filepath', ''),
                file_type=full_data.get('file_type', ''),
                company_name=full_data.get('company_name', ''),
                proposal_title=full_data.get('proposal_title', ''),
                date=full_data.get('date', ''),
                total_amount=full_data.get('total_amount'),
                total_raw=full_data.get('total_raw', ''),
                currency=full_data.get('currency', 'USD'),
                page_count=full_data.get('page_count', 0),
                extraction_notes=full_data.get('extraction_notes', []),
            )

            for td in full_data.get('tables', []):
                t = ExtractedTable(
                    headers=td.get('headers', []),
                    rows=td.get('rows', []),
                    source=td.get('source', ''),
                    table_index=td.get('table_index', 0),
                    has_financial_data=td.get('has_financial_data', False),
                    total_row_index=td.get('total_row_index'),
                )
                p.tables.append(t)

            for lid in full_data.get('line_items', []):
                li = LineItem(
                    description=lid.get('description', ''),
                    amount=lid.get('amount'),
                    amount_raw=lid.get('amount_raw', ''),
                    quantity=lid.get('quantity'),
                    unit_price=lid.get('unit_price'),
                    unit=lid.get('unit', ''),
                    category=lid.get('category', ''),
                    row_index=lid.get('row_index', 0),
                    table_index=lid.get('table_index', 0),
                    source_sheet=lid.get('source_sheet', ''),
                    confidence=lid.get('confidence', 1.0),
                )
                p.line_items.append(li)

            proposals.append(p)
            proposal_db_ids.append(prop_summary['id'])

        if len(proposals) < 2:
            return jsonify({
                'success': False,
                'error': {'message': 'Could not load enough proposals with valid data'}
            }), 400

        result = compare_proposals(proposals)
        result_dict = result.to_dict()

        # Save comparison
        save_comparison(project_id, proposal_db_ids, result_dict)

        return jsonify({'success': True, 'data': result_dict})

    except Exception as e:
        logger.error(f'Project compare error: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': {'message': f'Compare error: {str(e)}', 'traceback': traceback.format_exc()}
        }), 500


# ──────────────────────────────────────────
# Metrics for Metrics & Analytics Dashboard
# ──────────────────────────────────────────

@pc_blueprint.route('/api/proposal-compare/metrics', methods=['GET'])
def proposal_metrics():
    """Get aggregated proposal comparison metrics for the M&A dashboard."""
    try:
        from .projects import get_proposal_metrics
        metrics = get_proposal_metrics()
        return jsonify({'success': True, 'data': metrics})
    except Exception as e:
        logger.error(f'Proposal metrics error: {e}', exc_info=True)
        return jsonify({'success': False, 'error': {'message': str(e)}}), 500


# ──────────────────────────────────────────
# Structure Analysis (privacy-safe parsing)
# ──────────────────────────────────────────

@pc_blueprint.route('/api/proposal-compare/analyze-structure', methods=['POST'])
def analyze_structure():
    """Upload a proposal and return a privacy-safe structural analysis.

    The analysis reveals table shapes, column patterns, category distribution,
    and extraction diagnostics WITHOUT exposing dollar amounts, company names,
    or proprietary line-item descriptions.

    Expects multipart/form-data with a single 'file' field.
    Returns JSON structure report, or downloads as .json file if ?download=1.
    """
    try:
        from .parser import SUPPORTED_EXTENSIONS
        from .structure_analyzer import analyze_proposal_structure

        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': {'message': 'No file provided. Use "file" field.'}
            }), 400

        f = request.files['file']
        if not f.filename:
            return jsonify({
                'success': False,
                'error': {'message': 'No file selected'}
            }), 400

        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in SUPPORTED_EXTENSIONS:
            return jsonify({
                'success': False,
                'error': {'message': f'Unsupported file type: {ext}. Supported: {", ".join(SUPPORTED_EXTENSIONS)}'}
            }), 400

        # Save to temp file
        import time as _time
        upload_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'temp', 'proposals'
        )
        os.makedirs(upload_dir, exist_ok=True)
        safe_name = f'{int(_time.time())}_struct_{f.filename}'
        temp_path = os.path.join(upload_dir, safe_name)

        try:
            f.save(temp_path)
        except Exception as save_err:
            return jsonify({
                'success': False,
                'error': {'message': f'Failed to save file: {save_err}'}
            }), 500

        # Run structure analysis
        try:
            analysis = analyze_proposal_structure(temp_path)
        finally:
            # Clean up temp file immediately — we don't need it for viewing
            try:
                os.unlink(temp_path)
            except Exception:
                pass

        # If download requested, return as downloadable JSON file
        if request.args.get('download') == '1':
            import io
            json_str = json.dumps(analysis, indent=2, ensure_ascii=False)
            buf = io.BytesIO(json_str.encode('utf-8'))
            buf.seek(0)
            base_name = os.path.splitext(f.filename)[0]
            return send_file(
                buf,
                mimetype='application/json',
                as_attachment=True,
                download_name=f'{base_name}_structure_analysis.json'
            )

        return jsonify({'success': True, 'data': analysis})

    except Exception as e:
        logger.error(f'Structure analysis error: {e}', exc_info=True)
        return jsonify({
            'success': False,
            'error': {'message': f'Analysis error: {str(e)}', 'traceback': traceback.format_exc()}
        }), 500
