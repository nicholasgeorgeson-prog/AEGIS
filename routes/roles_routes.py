"""Roles Blueprint - Contains all role-related API routes."""

import json
import sqlite3
import socket
import uuid
from datetime import datetime, timezone
from pathlib import Path

from flask import Blueprint, jsonify, request, make_response, g, session
from routes._shared import (
    require_csrf,
    handle_api_errors,
    config,
    logger,
    SessionManager,
    ValidationError,
    sanitize_filename
)
import routes._shared as _shared


def get_scan_history_db():
    """Get scan history database instance."""
    from scan_history import get_scan_history_db as _get_db
    return _get_db()


# Create the roles blueprint
roles_bp = Blueprint('roles', __name__)

@roles_bp.route('/api/roles/aggregated', methods=['GET'])
@handle_api_errors
def get_aggregated_roles():
    """Get roles aggregated across all scanned documents."""
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role aggregation not available'})
    else:
        include_deliverables = request.args.get('include_deliverables', 'false').lower() == 'true'
        db = get_scan_history_db()
        roles = db.get_all_roles(include_deliverables)
        return jsonify({'success': True, 'data': roles})
@roles_bp.route('/api/roles/context', methods=['GET'])
@handle_api_errors
def get_role_context():
    """Get detailed context for a specific role including all occurrences.

    v3.1.9: Added for Role Source Viewer to show where roles were extracted.

    Query parameters:
        role: Name of the role to look up (required)

    Returns:
        Role details with occurrences across all documents
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role context not available'})
    else:
        role_name = request.args.get('role', '').strip()
        if not role_name:
            return (jsonify({'success': False, 'error': 'Role name is required'}), 400)
        else:
            db = get_scan_history_db()
            context = db.get_role_context(role_name)
            if not context:
                return (jsonify({'success': False, 'error': f'Role \"{role_name}\" not found'}), 404)
            else:
                return jsonify({'success': True, **context})


@roles_bp.route('/api/roles/responsibility', methods=['PUT'])
@require_csrf
@handle_api_errors
def update_responsibility_statement():
    """v4.8.4: Update an individual responsibility statement.

    Request body:
        role_name: Name of the role (required)
        document: Filename of the source document (required)
        statement_index: 0-based index in the document's responsibilities array (required)
        updates: Dict of fields to update - text, action_type, section, review_status, notes
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Database not available'})

    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Request body required'}), 400

    role_name = data.get('role_name', '').strip()
    document = data.get('document', '').strip()
    statement_index = data.get('statement_index')
    updates = data.get('updates', {})

    if not role_name:
        return jsonify({'success': False, 'error': 'role_name is required'}), 400
    if not document:
        return jsonify({'success': False, 'error': 'document is required'}), 400
    if statement_index is None or not isinstance(statement_index, int):
        return jsonify({'success': False, 'error': 'statement_index (integer) is required'}), 400
    if not updates:
        return jsonify({'success': False, 'error': 'updates object is required'}), 400

    db = get_scan_history_db()
    success = db.update_responsibility_statement(role_name, document, statement_index, updates)

    if success:
        return jsonify({'success': True, 'message': 'Statement updated'})
    else:
        return jsonify({'success': False, 'error': 'Statement not found or could not be updated'}), 404


@roles_bp.route('/api/roles/all-statements', methods=['GET'])
@handle_api_errors
def get_all_role_statements():
    """v4.9.5: Get all responsibility statements across all roles for mass review.

    Query parameters:
        review_status: Filter by review status (pending, reviewed, rejected, unreviewed)
        document: Filter by document filename (partial match)
        role: Filter by role name (partial match)
        search: Search text within statements
        flagged_only: If 'true', only return flagged statements
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Database not available'})

    filters = {}
    if request.args.get('review_status'):
        filters['review_status'] = request.args.get('review_status')
    if request.args.get('document'):
        filters['document'] = request.args.get('document')
    if request.args.get('role'):
        filters['role'] = request.args.get('role')
    if request.args.get('search'):
        filters['search'] = request.args.get('search')
    if request.args.get('flagged_only', '').lower() == 'true':
        filters['flagged_only'] = True

    db = get_scan_history_db()
    result = db.get_all_role_statements(filters)
    return jsonify({'success': True, **result})


@roles_bp.route('/api/roles/bulk-delete-statements', methods=['POST'])
@require_csrf
@handle_api_errors
def bulk_delete_statements():
    """v4.9.5: Bulk delete responsibility statements.

    Request body:
        deletions: Array of {role_name, document, statement_index}
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Database not available'})

    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Request body required'}), 400

    deletions = data.get('deletions', [])
    if not deletions:
        return jsonify({'success': False, 'error': 'No deletions specified'}), 400
    if len(deletions) > 1000:
        return jsonify({'success': False, 'error': 'Maximum 1000 deletions per request'}), 400

    db = get_scan_history_db()
    deleted = db.bulk_delete_role_statements(deletions)
    return jsonify({'success': True, 'deleted': deleted})


@roles_bp.route('/api/roles/bulk-update-statements', methods=['PUT'])
@require_csrf
@handle_api_errors
def bulk_update_statements():
    """v4.9.5: Bulk update responsibility statements.

    Request body:
        statements: Array of {role_name, document, statement_index}
        updates: Dict of fields to apply to all selected statements
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Database not available'})

    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'Request body required'}), 400

    statements = data.get('statements', [])
    updates = data.get('updates', {})
    if not statements:
        return jsonify({'success': False, 'error': 'No statements specified'}), 400
    if not updates:
        return jsonify({'success': False, 'error': 'No updates specified'}), 400
    if len(statements) > 1000:
        return jsonify({'success': False, 'error': 'Maximum 1000 statements per request'}), 400

    db = get_scan_history_db()
    updated = 0
    for stmt in statements:
        role_name = stmt.get('role_name', '')
        document = stmt.get('document', '')
        stmt_idx = stmt.get('statement_index')
        if role_name and document and stmt_idx is not None:
            if db.update_responsibility_statement(role_name, document, stmt_idx, updates):
                updated += 1

    return jsonify({'success': True, 'updated': updated})


@roles_bp.route('/api/roles/matrix', methods=['GET'])
@handle_api_errors
def get_role_document_matrix():
    """Get role-document relationship matrix for visualization."""
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role matrix not available'})
    else:
        db = get_scan_history_db()
        matrix = db.get_role_document_matrix()
        return jsonify({'success': True, 'data': matrix})
@roles_bp.route('/api/roles/matrix/export', methods=['POST'])
@require_csrf
@handle_api_errors
def export_role_document_matrix():
    """Export role-document matrix as XLSX or CSV.
    v4.7.0: Proper Excel export using openpyxl.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role matrix not available'})
    data = request.get_json() or {}
    fmt = data.get('format', 'xlsx')
    db = get_scan_history_db()
    matrix = db.get_role_document_matrix()
    roles = matrix.get('roles', {})
    documents = matrix.get('documents', [])
    if fmt == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from io import BytesIO
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Role-Document Matrix'
            # Header row
            headers = ['Role'] + [d.get('filename', d) if isinstance(d, dict) else str(d) for d in documents]
            header_fill = PatternFill(start_color='D6A84A', end_color='D6A84A', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                cell.border = thin_border
            # Data rows
            for row_idx, (role_name, role_data) in enumerate(sorted(roles.items()), 2):
                ws.cell(row=row_idx, column=1, value=role_name).border = thin_border
                doc_map = {}
                if isinstance(role_data, dict):
                    for doc in role_data.get('documents', []):
                        doc_name = doc.get('filename', doc) if isinstance(doc, dict) else str(doc)
                        doc_map[doc_name] = doc.get('mentions', 1) if isinstance(doc, dict) else 1
                for col_idx, doc in enumerate(documents, 2):
                    doc_name = doc.get('filename', doc) if isinstance(doc, dict) else str(doc)
                    count = doc_map.get(doc_name, 0)
                    cell = ws.cell(row=row_idx, column=col_idx, value=count if count else '')
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    if count:
                        cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            # Auto-fit column widths
            ws.column_dimensions['A'].width = 30
            for col_idx in range(2, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 15
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            response = make_response(buf.getvalue())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = f'attachment; filename=AEGIS_Role_Document_Matrix_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
            return response
        except ImportError:
            return jsonify({'success': False, 'error': 'openpyxl not available for Excel export'}), 500
    return jsonify({'success': False, 'error': f'Unsupported format: {fmt}'}), 400

@roles_bp.route('/api/roles/raci', methods=['GET'])
@handle_api_errors
def get_raci_matrix():
    """
    Get enhanced RACI matrix computed from stored responsibilities.

    v3.1.10: New endpoint for comprehensive RACI analysis.

    RACI Categories:
    - R (Responsible): Performs the work
    - A (Accountable): Approves/authorizes
    - C (Consulted): Provides input
    - I (Informed): Kept in loop

    Query parameters:
        include_documents: Include document breakdown (default true)

    Returns:
        roles: Dict of role names to RACI data
        summary: Aggregate statistics and distribution
        documents: Document ID to filename mapping
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'RACI matrix not available'})
    else:
        include_documents = request.args.get('include_documents', 'true').lower() == 'true'
        db = get_scan_history_db()
        raci_data = db.get_raci_matrix(include_documents=include_documents)
        return jsonify({'success': True, 'data': raci_data})

@roles_bp.route('/api/roles/verify', methods=['POST'])
@require_csrf
@handle_api_errors
def verify_roles():
    """
    Verify roles against stored document text and optionally cleanup unverified.

    v3.1.10: Prevents hallucinated roles from persisting.

    Request body:
        dry_run: If true (default), only report without deleting

    Returns:
        Verification results with counts of verified/unverified roles
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role verification not available'})
    else:
        data = request.get_json() or {}
        dry_run = data.get('dry_run', True)
        db = get_scan_history_db()
        result = db.verify_and_cleanup_roles(dry_run=dry_run)
        return jsonify({'success': True, 'data': result})
@roles_bp.route('/api/roles/extraction-mode', methods=['GET', 'POST'])
@handle_api_errors
def extraction_mode():
    """
    v3.4.0: Get or set the role extraction mode.

    GET: Returns current extraction mode
    POST: Sets extraction mode
        Request body: {\"mode\": \"strict\" | \"discovery\"}

        - strict: Whitelist-only extraction with 100% accuracy guarantee.
                  Only extracts roles that exist in KNOWN_ROLES and appear
                  verbatim in the document. No false positives.
        - discovery: Pattern-based extraction for maximum recall.
                     May include potential new roles that need verification.
    """
    try:
        from role_integration import RoleIntegration
        if request.method == 'GET':
            integration = RoleIntegration()
            return jsonify({'success': True, 'mode': integration.get_extraction_mode(), 'available_modes': ['discovery', 'strict'], 'descriptions': {'discovery': 'Pattern-based extraction (may include unverified roles)', 'strict': 'Whitelist-only (100% accuracy, verifiable in source)'}})
        else:
            data = request.get_json() or {}
            mode = data.get('mode', 'discovery').lower()
            if mode not in ['discovery', 'strict']:
                return jsonify({'success': False, 'error': f'Invalid mode: {mode}. Use \'discovery\' or \'strict\'.'})
            else:
                integration = RoleIntegration(extraction_mode=mode)
                return jsonify({'success': True, 'mode': integration.get_extraction_mode(), 'message': f'Extraction mode set to {mode.upper()}'})
    except ImportError as e:
        return jsonify({'success': False, 'error': f'Role extraction not available: {e}'})
@roles_bp.route('/api/roles/graph', methods=['GET'])
@handle_api_errors
def get_role_graph():
    """
    Get graph data for D3.js visualization of role-document relationships.
    
    Query parameters:
    - max_nodes: Maximum nodes to return (default 100, max 500)
    - min_weight: Minimum edge weight to include (default 1)
    - use_cache: Whether to use cached data (default true)
    
    Returns:
    - nodes: Array of role and document nodes with stable IDs
    - links: Array of edges with weights
    - role_counts: Aggregate stats per role
    - doc_counts: Aggregate stats per document
    - meta: Query metadata
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role graph not available', 'data': {'nodes': [], 'links': [], 'role_counts': {}, 'doc_counts': {}}})
    else:
        try:
            max_nodes = min(int(request.args.get('max_nodes', 100)), 500)
            min_weight = max(int(request.args.get('min_weight', 1)), 1)
            use_cache = request.args.get('use_cache', 'true').lower()!= 'false'
        except ValueError:
            max_nodes = 100
            min_weight = 1
            use_cache = True
        db = get_scan_history_db()
        session_id = session.get('session_id', 'default')
        sess_data = SessionManager.get(session_id)
        file_hash = ''
        if sess_data and sess_data.get('current_file'):
            import hashlib
            try:
                with open(sess_data['current_file'], 'rb') as f:
                    file_hash = hashlib.md5(f.read(10000)).hexdigest()
            except Exception as e:
                logger.warning(f'Error computing file hash: {e}')
        if use_cache and file_hash:
            from scan_history import get_cached_graph
            graph_data = get_cached_graph(session_id, file_hash, db, max_nodes, min_weight)
        else:
            graph_data = db.get_role_graph_data(max_nodes, min_weight)
        return jsonify({'success': True, 'data': graph_data})


@roles_bp.route('/api/roles/graph-export-html', methods=['GET'])
@handle_api_errors
def export_graph_html():
    """
    v5.9.16: Export the role relationship graph as a standalone interactive HTML file.
    Includes embedded D3.js, pan/zoom, search, and filter by function/role type.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Graph data not available'})

    db = get_scan_history_db()
    graph_data = db.get_role_graph_data(max_nodes=500, min_weight=1)

    # Get role dictionary for enrichment
    dictionary = db.get_role_dictionary(include_inactive=False)
    role_lookup = {}
    for r in dictionary:
        rn = r.get('role_name', '').lower().strip()
        role_lookup[rn] = {
            'role_type': r.get('role_type', ''),
            'category': r.get('category', ''),
            'is_deliverable': r.get('is_deliverable', False),
            'org_group': r.get('org_group', ''),
            'description': r.get('description', '')
        }

    # Get function tags
    with db.connection() as (conn, cursor):
        cursor.execute('''
            SELECT rft.role_name, rft.function_code,
                   fc.name as function_name, fc.color as function_color
            FROM role_function_tags rft
            LEFT JOIN function_categories fc ON fc.code = rft.function_code
        ''')
        tag_rows = cursor.fetchall()
        tag_lookup = {}
        for tr in tag_rows:
            rn = tr['role_name'].lower().strip()
            if rn not in tag_lookup:
                tag_lookup[rn] = []
            tag_lookup[rn].append({
                'code': tr['function_code'],
                'name': tr['function_name'] or tr['function_code'],
                'color': tr['function_color'] or '#3b82f6'
            })

    # Enrich nodes with role metadata
    node_name_to_id = {}
    for node in graph_data.get('nodes', []):
        if node.get('type') == 'role':
            rn = node.get('name', '').lower().strip()
            node_name_to_id[rn] = node.get('id')
            meta = role_lookup.get(rn, {})
            node['role_type'] = meta.get('role_type', '')
            node['category'] = meta.get('category', '')
            node['is_deliverable'] = meta.get('is_deliverable', False)
            node['org_group'] = meta.get('org_group', '')
            node['description'] = meta.get('description', '')
            node['function_tags'] = tag_lookup.get(rn, [])

    # v5.9.23: Add directional role_relationships to links
    try:
        relationships = db.get_role_relationships()
        existing_links = graph_data.get('links', [])
        for rel in relationships:
            src_name = (rel.get('source_role_name', '') or '').lower().strip()
            tgt_name = (rel.get('target_role_name', '') or '').lower().strip()
            src_id = node_name_to_id.get(src_name)
            tgt_id = node_name_to_id.get(tgt_name)
            if src_id and tgt_id:
                existing_links.append({
                    'source': src_id,
                    'target': tgt_id,
                    'link_type': 'relationship',
                    'relationship_type': rel.get('relationship_type', 'related'),
                    'weight': 2
                })
        graph_data['links'] = existing_links
    except Exception as rel_err:
        current_app.logger.warning(f'Could not load role_relationships for graph export: {rel_err}')

    from graph_export_html import generate_graph_html
    from config_logging import get_version

    html_content = generate_graph_html(
        graph_data=graph_data,
        metadata={
            'version': get_version(),
            'export_date': datetime.now(timezone.utc).isoformat(),
            'hostname': socket.gethostname()
        }
    )

    response = make_response(html_content)
    response.headers['Content-Type'] = 'text/html; charset=utf-8'
    date_str = datetime.now().strftime('%Y-%m-%d')
    response.headers['Content-Disposition'] = f'attachment; filename=AEGIS_Graph_{date_str}.html'
    return response


@roles_bp.route('/api/roles/adjudicate', methods=['POST'])
@require_csrf
@handle_api_errors
def adjudicate_role():
    """
    Adjudicate a role - confirm, mark as deliverable, or reject.

    This endpoint handles role adjudication from the Role Source Viewer:
    - Confirmed roles are added to the dictionary with source=\'adjudication\'
    - Deliverables are added with is_deliverable=True
    - Rejected roles are added with is_active=False (excluded from future extraction)

    Request body:
    - role_name: Required - the role name to adjudicate
    - action: Required - \'confirmed\', \'deliverable\', or \'rejected\'
    - category: Optional - role category
    - notes: Optional - adjudication notes
    - source_document: Optional - document where role was found
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        data = request.get_json() or {}
        role_name = data.get('role_name')
        action = data.get('action')
        if not role_name:
            raise ValidationError('role_name is required')
        else:
            if action not in ['confirmed', 'deliverable', 'rejected']:
                raise ValidationError('action must be \'confirmed\', \'deliverable\', or \'rejected\'')
            else:
                db = get_scan_history_db()
                existing = db.get_role_by_name(role_name)
                if existing:
                    update_data = {'is_active': action!= 'rejected', 'is_deliverable': action == 'deliverable', 'category': data.get('category', existing.get('category', 'Role')), 'notes': data.get('notes', existing.get('notes', '')), 'source': 'adjudication'}
                    result = db.update_role_in_dictionary(existing['id'], updated_by='adjudication', **update_data)
                    if result.get('success'):
                        function_tags = data.get('function_tags', [])
                        if function_tags:
                            with db.connection() as (conn, cursor):
                                for tag_code in function_tags:
                                    try:
                                        cursor.execute('\n                            INSERT OR IGNORE INTO role_function_tags (role_id, role_name, function_code, assigned_by)\n                            VALUES (?, ?, ?, \'adjudication\')\n                        ', (existing['id'], role_name, tag_code.strip().upper()))
                                    except Exception as e:
                                        logger.warning(f'Error inserting role function tag: {e}')
                        return jsonify({'success': True, 'message': f'Role \'{role_name}\' updated - {action}', 'role_id': existing['id'], 'action': action, 'is_new': False, 'function_tags': function_tags})
                    else:
                        return jsonify({'success': False, 'error': result.get('error', 'Failed to update role')})
                else:
                    result = db.add_role_to_dictionary(role_name=role_name, source='adjudication', category=data.get('category', 'Role'), source_document=data.get('source_document'), is_deliverable=action == 'deliverable', is_active=action!= 'rejected', created_by='adjudication', notes=data.get('notes', f'Adjudicated as {action}'))
                    if result.get('success'):
                        role_id = result.get('role_id') or result.get('id')
                        function_tags = data.get('function_tags', [])
                        if function_tags and role_id:
                                with db.connection() as (conn, cursor):
                                    for tag_code in function_tags:
                                        try:
                                            cursor.execute('\n                            INSERT OR IGNORE INTO role_function_tags (role_id, role_name, function_code, assigned_by)\n                            VALUES (?, ?, ?, \'adjudication\')\n                        ', (role_id, role_name, tag_code.strip().upper()))
                                        except Exception:
                                            pass
                        return jsonify({'success': True, 'message': f'Role \'{role_name}\' added - {action}', 'role_id': role_id, 'action': action, 'is_new': True, 'function_tags': function_tags})
                return jsonify({'success': False, 'error': result.get('error', 'Failed to adjudicate role')})
@roles_bp.route('/api/roles/rename', methods=['POST'])
@require_csrf
@handle_api_errors
def rename_role():
    """
    Rename a role across the system.

    Updates role_dictionary.role_name and role_function_tags.role_name.

    Request body:
    - old_name: Required - current role name
    - new_name: Required - new role name
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        data = request.get_json() or {}
        old_name = data.get('old_name', '').strip()
        new_name = data.get('new_name', '').strip()
        if not old_name or not new_name:
            raise ValidationError('Both old_name and new_name are required')
        else:
            if old_name.lower() == new_name.lower():
                return jsonify({'success': True, 'message': 'No change needed'})
            else:
                db = get_scan_history_db()
                conflict = db.get_role_by_name(new_name)
                existing = db.get_role_by_name(old_name)
                if existing:
                    if conflict and conflict['id']!= existing['id']:
                        return jsonify({'success': False, 'error': f'Role \"{new_name}\" already exists in dictionary'})
                    else:
                        role_id = existing['id']
                        result = db.update_role_in_dictionary(role_id, updated_by='rename', role_name=new_name)
                        if not result.get('success'):
                            return jsonify({'success': False, 'error': 'Failed to update role dictionary'})
                        else:
                            try:
                                with db.connection() as (conn, cursor):
                                    cursor.execute('UPDATE role_function_tags SET role_name = ? WHERE role_name = ?', (new_name, old_name))
                            except Exception as e:
                                logger.warning(f'Failed to update role_function_tags for rename: {e}')
                            return jsonify({'success': True, 'message': f'Role renamed from \"{old_name}\" to \"{new_name}\"', 'old_name': old_name, 'new_name': new_name, 'role_id': role_id})
                else:
                    if conflict:
                        return jsonify({'success': False, 'error': f'Role \"{new_name}\" already exists in dictionary'})
                    else:
                        result = db.add_role_to_dictionary(new_name, source='rename', category='Role', created_by='rename')
                        if result.get('success') or result.get('id'):
                            role_id = result.get('id') or result.get('role_id')
                            return jsonify({'success': True, 'message': f'Role \"{old_name}\" added to dictionary as \"{new_name}\"', 'old_name': old_name, 'new_name': new_name, 'role_id': role_id, 'is_new': True})
                        else:
                            return jsonify({'success': False, 'error': result.get('error', 'Failed to add role to dictionary')})
@roles_bp.route('/api/roles/adjudication-status', methods=['GET'])
@handle_api_errors
def get_role_adjudication_status():
    """
    Get the adjudication status for a role.

    Query parameters:
    - role_name: The role name to check

    Returns:
    - status: \'pending\', \'confirmed\', \'deliverable\', or \'rejected\'
    - category: The role category
    - notes: Any adjudication notes
    - adjudicated_at: When it was adjudicated (if applicable)
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        role_name = request.args.get('role_name')
        if not role_name:
            raise ValidationError('role_name is required')
        else:
            db = get_scan_history_db()
            existing = db.get_role_by_name(role_name)
            if not existing:
                return jsonify({'success': True, 'status': 'pending', 'category': 'Role', 'notes': '', 'adjudicated_at': None})
            else:
                if not existing.get('is_active', True):
                    status = 'rejected'
                else:
                    if existing.get('is_deliverable', False):
                        status = 'deliverable'
                    else:
                        status = 'confirmed'
                return jsonify({'success': True, 'status': status, 'category': existing.get('category', 'Role'), 'notes': existing.get('notes', ''), 'adjudicated_at': existing.get('updated_at') or existing.get('created_at'), 'role_id': existing.get('id')})
@roles_bp.route('/api/roles/auto-adjudicate', methods=['POST'])
@require_csrf
@handle_api_errors
def auto_adjudicate_roles():
    """
    Auto-classify all pending roles based on dictionary, patterns, and confidence.

    Optional request body:
    - apply: bool - If true, auto-apply suggestions above threshold (default false)
    - threshold: float - Confidence threshold for auto-apply (default 0.85)

    Returns suggestions with recommended status for each role.
    """
    import re
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Scan history not available'})
    else:
        data = request.get_json() or {}
        apply_suggestions = data.get('apply', False)
        threshold = data.get('threshold', 0.85)
        db = get_scan_history_db()
        roles = db.get_all_roles()
        deliverable_patterns = [re.compile('\\b(document|report|plan|specification|analysis|review|audit|assessment)\\b', re.I), re.compile('\\b(drawing|schematic|diagram|model|prototype)\\b', re.I), re.compile('\\b(database|repository|archive|library)\\b', re.I), re.compile('\\b(schedule|timeline|roadmap|milestone)\\b', re.I), re.compile('\\b(budget|estimate|proposal|contract)\\b', re.I), re.compile('\\b(test|verification|validation)\\s+(report|results|data)\\b', re.I), re.compile('\\b(requirements|interface|design)\\s+(document|spec)\\b', re.I), re.compile('\\bICD\\b|\\bSRS\\b|\\bSDD\\b|\\bCDRL\\b|\\bDID\\b', re.I)]
        role_patterns = [re.compile('\\b(engineer|manager|lead|director|officer|specialist|analyst)\\b', re.I), re.compile('\\b(coordinator|administrator|supervisor|inspector|reviewer)\\b', re.I), re.compile('\\b(team|group|committee|board|panel|council)\\b', re.I), re.compile('\\b(technician|operator|programmer|designer|planner)\\b', re.I), re.compile('\\b(chief|head|principal|senior|junior)\\b', re.I)]
        suggestions = []
        auto_applied = 0
        for role in roles:
            role_name = role.get('role_name', '')
            existing = db.get_role_by_name(role_name)
            if existing:
                if not existing.get('is_active', True):
                    status = 'rejected'
                else:
                    if existing.get('is_deliverable', False):
                        status = 'deliverable'
                    else:
                        status = 'confirmed'
                suggestions.append({'role_name': role_name, 'suggested_status': status, 'confidence': 0.99, 'reason': f'Already in dictionary as {status}', 'source': 'dictionary', 'already_adjudicated': True})
            else:
                confidence = 0.5
                suggested_status = 'pending'
                reason = 'No strong pattern match'
                is_deliverable = any((p.search(role_name) for p in deliverable_patterns))
                if is_deliverable:
                    suggested_status = 'deliverable'
                    confidence = 0.8
                    reason = 'Name matches deliverable/artifact pattern'
                is_role = any((p.search(role_name) for p in role_patterns))
                if is_role and (not is_deliverable):
                        suggested_status = 'confirmed'
                        confidence = 0.85
                        reason = 'Name matches role/title pattern'
                doc_count = role.get('unique_document_count', role.get('document_count', 1))
                mentions = role.get('total_mentions', 1)
                resp_count = role.get('responsibility_count', 0)
                if doc_count >= 3:
                    confidence = min(confidence + 0.1, 0.99)
                    reason += f' (found in {doc_count} documents)'
                if mentions >= 5:
                    confidence = min(confidence + 0.05, 0.99)
                if resp_count >= 2:
                    confidence = min(confidence + 0.05, 0.99)
                    if suggested_status == 'pending':
                        suggested_status = 'confirmed'
                        reason = f'Has {resp_count} responsibilities assigned'
                suggestions.append({'role_name': role_name, 'suggested_status': suggested_status, 'confidence': round(confidence, 2), 'reason': reason, 'source': 'pattern', 'already_adjudicated': False, 'document_count': doc_count, 'mention_count': mentions, 'responsibility_count': resp_count})
        if apply_suggestions:
            to_apply = [s for s in suggestions if not s.get('already_adjudicated') and s['confidence'] >= threshold and s['suggested_status'] != 'pending']
            if to_apply:
                decisions = [{'role_name': s['role_name'], 'action': s['suggested_status'], 'notes': f"Auto-classified: {s['reason']} (confidence: {s['confidence']})"} for s in to_apply]
                result = db.batch_adjudicate(decisions)
                auto_applied = result.get('processed', 0)
        return jsonify({'success': True, 'suggestions': suggestions, 'auto_applied': auto_applied, 'total': len(suggestions), 'pending_count': sum(1 for s in suggestions if s['suggested_status'] == 'pending'), 'actionable_count': sum(1 for s in suggestions if not s.get('already_adjudicated') and s['suggested_status'] != 'pending')})
@roles_bp.route('/api/roles/adjudicate/batch', methods=['POST'])
@require_csrf
@handle_api_errors
def batch_adjudicate_roles():
    """
    Batch adjudicate multiple roles in a single request.

    Request body:
    - decisions: Array of {role_name, action, category?, notes?, function_tags?[]}

    Returns processed count and any errors.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Scan history not available'})
    else:
        data = request.get_json() or {}
        decisions = data.get('decisions', [])
        if not decisions or not isinstance(decisions, list):
            raise ValidationError('decisions array is required')
        else:
            if len(decisions) > 500:
                raise ValidationError('Maximum 500 decisions per batch')
            else:
                db = get_scan_history_db()
                result = db.batch_adjudicate(decisions)
                return jsonify({'success': True, 'processed': result.get('processed', 0), 'total': result.get('total', 0), 'errors': result.get('errors', []), 'results': result.get('results', [])})
@roles_bp.route('/api/roles/adjudication-summary', methods=['GET'])
@handle_api_errors
def get_adjudication_summary():
    """Get adjudication statistics and recent activity."""
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Scan history not available'})
    else:
        db = get_scan_history_db()
        summary = db.get_adjudication_summary()
        return jsonify({'success': True, 'data': summary})
@roles_bp.route('/api/roles/update-category', methods=['POST'])
@require_csrf
@handle_api_errors
def update_role_category():
    """
    Update the category for a role.

    Request body:
    - role_name: Required
    - category: Required - new category value
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        data = request.get_json() or {}
        role_name = data.get('role_name')
        category = data.get('category')
        if not role_name:
            raise ValidationError('role_name is required')
        else:
            if not category:
                raise ValidationError('category is required')
            else:
                db = get_scan_history_db()
                existing = db.get_role_by_name(role_name)
                if existing:
                    result = db.update_role_in_dictionary(existing['id'], updated_by='category_update', category=category)
                    return jsonify(result)
                else:
                    result = db.add_role_to_dictionary(role_name=role_name, source='category_update', category=category, created_by='user')
                    return jsonify(result)
@roles_bp.route('/api/roles/dictionary', methods=['GET'])
@handle_api_errors
def get_role_dictionary():
    """
    Get all roles from the role dictionary.
    
    Query parameters:
    - include_inactive: Include deactivated roles (default false)
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
        db = get_scan_history_db()
        roles = db.get_role_dictionary(include_inactive)
        return jsonify({'success': True, 'data': {'roles': roles, 'total': len(roles)}})


@roles_bp.route('/api/roles/dictionary/stats', methods=['GET'])
@handle_api_errors
def get_role_dictionary_stats():
    """
    Lightweight stats for the landing page — avoids loading full role data.
    Returns counts only: total, adjudicated, deliverable, by_category.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    db = get_scan_history_db()
    with db.connection() as (conn, cursor):
        cursor.execute('''
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN role_disposition IS NOT NULL AND role_disposition != '' AND role_disposition != 'pending' THEN 1 ELSE 0 END) as adjudicated,
                SUM(CASE WHEN is_deliverable = 1 THEN 1 ELSE 0 END) as deliverable
            FROM role_dictionary WHERE is_active = 1
        ''')
        row = cursor.fetchone()
        return jsonify({'success': True, 'data': {
            'total': row[0] or 0,
            'adjudicated': row[1] or 0,
            'deliverable': row[2] or 0
        }})


@roles_bp.route('/api/roles/dictionary', methods=['POST'])
@require_csrf
@handle_api_errors
def add_role_to_dictionary():
    """
    Add a new role to the dictionary.
    
    Request body:
    - role_name: Required
    - source: Source of role (\'manual\', \'adjudication\', \'upload\')
    - category: Optional category
    - aliases: Optional list of aliases
    - description: Optional description
    - is_deliverable: Optional boolean
    - notes: Optional notes
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        data = request.get_json() or {}
        role_name = data.get('role_name')
        if not role_name:
            raise ValidationError('role_name is required')
        else:
            source = data.get('source', 'manual')
            db = get_scan_history_db()
            result = db.add_role_to_dictionary(role_name=role_name, source=source, category=data.get('category'), aliases=data.get('aliases', []), description=data.get('description'), source_document=data.get('source_document'), is_deliverable=data.get('is_deliverable', False), created_by=data.get('created_by', 'user'), notes=data.get('notes'))
            return jsonify(result)
@roles_bp.route('/api/roles/dictionary/<int:role_id>', methods=['PUT'])
@require_csrf
@handle_api_errors
def update_role_in_dictionary(role_id):
    """Update an existing role in the dictionary."""
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        data = request.get_json() or {}
        updated_by = data.pop('updated_by', 'user')
        db = get_scan_history_db()
        result = db.update_role_in_dictionary(role_id, updated_by=updated_by, **data)
        return jsonify(result)
@roles_bp.route('/api/roles/dictionary/<int:role_id>', methods=['DELETE'])
@require_csrf
@handle_api_errors
def delete_role_from_dictionary(role_id):
    """Delete (deactivate) a role from the dictionary."""
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        hard_delete = request.args.get('hard', 'false').lower() == 'true'
        db = get_scan_history_db()
        result = db.delete_role_from_dictionary(role_id, soft_delete=not hard_delete)
        return jsonify(result)
@roles_bp.route('/api/roles/dictionary/import', methods=['POST'])
@require_csrf
@handle_api_errors
def import_roles_to_dictionary():
    """
    Bulk import roles to the dictionary.
    
    Accepts either:
    - JSON body with \'roles\' array
    - File upload (CSV or JSON)
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        db = get_scan_history_db()
        if 'file' in request.files:
            file = request.files['file']
            if not file.filename:
                raise ValidationError('No file selected')
            else:
                filename = sanitize_filename(file.filename)
                content = file.read().decode('utf-8')
                if filename.endswith('.json'):
                    import json as json_module
                    data = json_module.loads(content)
                    roles = data if isinstance(data, list) else data.get('roles', [])
                else:
                    if filename.endswith('.csv'):
                        import csv
                        import io
                        reader = csv.DictReader(io.StringIO(content))
                        roles = list(reader)
                    else:
                        raise ValidationError('Unsupported file type. Use .json or .csv')
                source = 'upload'
                source_document = filename
                created_by = request.form.get('created_by', 'file_import')
        else:
            data = request.get_json() or {}
            roles = data.get('roles', [])
            source = data.get('source', 'manual')
            source_document = data.get('source_document')
            created_by = data.get('created_by', 'user')
        if not roles:
            raise ValidationError('No roles provided')
        else:
            result = db.import_roles_to_dictionary(roles=roles, source=source, source_document=source_document, created_by=created_by)
            return jsonify({'success': True, 'data': result})
@roles_bp.route('/api/roles/dictionary/seed', methods=['POST'])
@require_csrf
@handle_api_errors
def seed_role_dictionary():
    """Seed the dictionary with built-in known roles."""
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        db = get_scan_history_db()
        result = db.seed_builtin_roles()
        return jsonify({'success': True, 'data': result})
@roles_bp.route('/api/roles/dictionary/import-excel', methods=['POST'])
@require_csrf
@handle_api_errors
def import_excel_to_dictionary():
    """
    v2.9.3 F01: Import roles from Excel file (process map export format).
    
    Parses Column J (Activity Resources) and Column L (Info 2/Description).
    Handles [S] prefix for Tools/Systems category.
    
    Returns preview data for user confirmation before actual import.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        if 'file' not in request.files:
            raise ValidationError('No file provided')
        else:
            file = request.files['file']
            if not file.filename:
                raise ValidationError('No file selected')
            else:
                filename = sanitize_filename(file.filename)
                if not filename.endswith(('.xlsx', '.xls')):
                    raise ValidationError('Please upload an Excel file (.xlsx or .xls)')
                else:
                    preview_mode = request.form.get('preview', 'true').lower() == 'true'
                    try:
                        import openpyxl
                    except ImportError:
                        return jsonify({'success': False, 'error': 'Excel support not available. Install openpyxl: pip install openpyxl'})
                    try:
                        temp_path = config.temp_dir / f'import_{uuid.uuid4().hex[:8]}_{filename}'
                        file.save(str(temp_path))
                        wb = openpyxl.load_workbook(str(temp_path), data_only=True)
                        ws = wb.active
                        roles_found = {}
                        tools_found = {}
                        col_j = 9
                        col_l = 11
                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                            if len(row) <= col_j:
                                continue
                            else:
                                resources_cell = row[col_j] if col_j < len(row) else None
                                desc_cell = row[col_l] if col_l < len(row) else None
                                if not resources_cell:
                                    continue
                                else:
                                    resource_str = str(resources_cell)
                                    description = str(desc_cell) if desc_cell else ''
                                    for resource in resource_str.split(';'):
                                        resource = resource.strip()
                                        if not resource:
                                            continue
                                        else:
                                            if resource.startswith('[S]'):
                                                tool_name = resource[3:].strip()
                                                if tool_name:
                                                    if tool_name not in tools_found:
                                                        tools_found[tool_name] = {'descriptions': set(), 'source_rows': []}
                                                    if description:
                                                        tools_found[tool_name]['descriptions'].add(description)
                                                    tools_found[tool_name]['source_rows'].append(row_idx)
                                            else:
                                                if resource not in roles_found:
                                                    roles_found[resource] = {'descriptions': set(), 'source_rows': []}
                                                if description:
                                                    roles_found[resource]['descriptions'].add(description)
                                                roles_found[resource]['source_rows'].append(row_idx)
                        temp_path.unlink(missing_ok=True)
                        wb.close()
                        human_roles = []
                        for name, data in sorted(roles_found.items()):
                            human_roles.append({'role_name': name, 'category': 'Unknown', 'description': '; '.join(sorted(data['descriptions']))[:500], 'source_rows': data['source_rows'][:5], 'occurrence_count': len(data['source_rows'])})
                        tools_systems = []
                        for name, data in sorted(tools_found.items()):
                            tools_systems.append({'role_name': name, 'category': 'Tools & Systems', 'description': '; '.join(sorted(data['descriptions']))[:500], 'source_rows': data['source_rows'][:5], 'occurrence_count': len(data['source_rows'])})
                        if preview_mode:
                            return jsonify({'success': True, 'preview': True, 'data': {'human_roles': human_roles, 'tools_systems': tools_systems, 'total_human_roles': len(human_roles), 'total_tools': len(tools_systems), 'source_file': filename}})
                        else:
                            selected_roles = request.form.getlist('selected_roles[]')
                            selected_tools = request.form.getlist('selected_tools[]')
                            db = get_scan_history_db()
                            roles_to_import = []
                            for role in human_roles:
                                if not selected_roles or role['role_name'] in selected_roles:
                                    roles_to_import.append({'role_name': role['role_name'], 'category': role.get('category', 'Unknown'), 'description': role.get('description', ''), 'aliases': []})
                            for tool in tools_systems:
                                if not selected_tools or tool['role_name'] in selected_tools:
                                    roles_to_import.append({'role_name': tool['role_name'], 'category': 'Tools & Systems', 'description': tool.get('description', ''), 'aliases': []})
                            result = db.import_roles_to_dictionary(roles=roles_to_import, source='excel_import', source_document=filename, created_by='excel_import')
                            return jsonify({'success': True, 'preview': False, 'data': result})
                    except Exception as e:
                        logger.exception(f'Error parsing Excel file: {e}')
                        return jsonify({'success': False, 'error': f'Error parsing Excel file: {str(e)}'})
@roles_bp.route('/api/roles/dictionary/export', methods=['GET'])
@handle_api_errors
def export_role_dictionary():
    """Export the role dictionary as JSON or CSV."""
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        format_type = request.args.get('format', 'json')
        include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
        db = get_scan_history_db()
        roles = db.get_role_dictionary(include_inactive)
        if format_type == 'csv':
            import csv
            import io
            output = io.StringIO()
            if roles:
                fieldnames = ['role_name', 'category', 'aliases', 'source', 'source_document', 'description', 'is_active', 'is_deliverable', 'created_at', 'created_by', 'updated_at', 'notes']
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for role in roles:
                    row = {k: role.get(k) for k in fieldnames}
                    row['aliases'] = ','.join(role.get('aliases', []))
                    writer.writerow(row)
            csv_content = output.getvalue()
            # Add UTF-8 BOM for proper Excel display on Windows
            csv_bytes = csv_content.encode('utf-8-sig')
            response = make_response(csv_bytes)
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = 'attachment; filename=role_dictionary.csv'
            return response
        else:
            return jsonify({'success': True, 'data': {'roles': roles, 'exported_at': datetime.now().isoformat(), 'total': len(roles)}})
@roles_bp.route('/api/roles/dictionary/import-sipoc', methods=['POST'])
@require_csrf
@handle_api_errors
def import_sipoc_to_dictionary():
    """
    v4.1.0: Import roles from Nimbus SIPOC Excel export.

    Parses 'Roles Hierarchy' map path rows, extracts role hierarchy,
    supervisory relationships, org metadata, and function tags.

    Query params:
      - confirm=true: Actually import (default is preview mode)

    Form data:
      - file: .xlsx file upload
      - clear_existing: 'true' to clear previous SIPOC import first
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    if 'file' not in request.files:
        raise ValidationError('No file provided')
    file = request.files['file']
    if not file.filename:
        raise ValidationError('No file selected')
    filename = sanitize_filename(file.filename)
    if not filename.endswith(('.xlsx', '.xls')):
        raise ValidationError('Please upload an Excel file (.xlsx or .xls)')
    confirm = request.args.get('confirm', 'false').lower() == 'true'
    clear_existing = request.form.get('clear_existing', 'false').lower() == 'true'
    try:
        from sipoc_parser import parse_sipoc_file
    except ImportError:
        return jsonify({'success': False, 'error': 'SIPOC parser not available. Ensure sipoc_parser.py is in the application directory.'})
    import uuid as _uuid
    temp_filename = f'sipoc_{_uuid.uuid4().hex[:8]}_{filename}'
    temp_path = config.temp_dir / temp_filename
    try:
        try:
            file.save(str(temp_path))
            parsed = parse_sipoc_file(str(temp_path))
            if not confirm:
                sample_roles = []
                for r in parsed['roles'][:20]:
                    func_tags = r.get('function_tags', [])
                    func_display = ', '.join(sorted(set(ft.get('parent_code') or ft.get('parent_name', '') for ft in func_tags if ft.get('parent_code') or ft.get('parent_name')))) if func_tags else ''
                    sample_roles.append({'role_name': r['role_name'], 'is_tool': r['is_tool'], 'category': r['category'], 'function_area': func_display, 'role_type': r.get('role_type', ''), 'role_disposition': r.get('role_disposition', ''), 'baselined': r.get('baselined', False), 'description': (r.get('description', '') or '')[:100], 'parents': r.get('parents', [])[:5], 'children': r.get('children', [])[:5], 'aliases': r.get('aliases', [])[:5], 'function_tags': func_tags[:5]})
                func_areas = set()
                for r in parsed['roles']:
                    for ft in r.get('function_tags', []):
                        p_name = ft.get('parent_name', '')
                        if p_name:
                            func_areas.add(p_name)
                        c_name = ft.get('child_name', '')
                        if c_name:
                            func_areas.add(c_name)
                categories = sorted(func_areas) if func_areas else sorted(set(r.get('org_group', '') for r in parsed['roles'] if r.get('org_group') and not r.get('is_tool')))
                rel_types = {}
                for rel in parsed['relationships']:
                    rt = rel.get('type', 'unknown')
                    rel_types[rt] = rel_types.get(rt, 0) + 1
                dispositions = {}
                for r in parsed['roles']:
                    d = r.get('role_disposition', '') or 'Not Set'
                    dispositions[d] = dispositions.get(d, 0) + 1
                role_types = {}
                for r in parsed['roles']:
                    rt = r.get('role_type', '') or 'Not Set'
                    role_types[rt] = role_types.get(rt, 0) + 1
                return jsonify({'success': True, 'preview': True, 'data': {'stats': parsed['stats'], 'sample_roles': sample_roles, 'categories': categories, 'relationship_types': rel_types, 'dispositions': dispositions, 'role_types': role_types, 'grouping_rows': [{'level': g.get('level', ''), 'title': g.get('diagram_title', ''), 'activity': g.get('activity_text', '')} for g in parsed.get('grouping_rows', [])[:30]]}})
            db = get_scan_history_db()
            if clear_existing:
                clear_result = db.clear_sipoc_import()
                logger.info(f'Cleared previous SIPOC import: {clear_result}')
            import_result = db.import_sipoc_roles(parsed, created_by='sipoc_import')
            return jsonify({'success': True, 'preview': False, 'data': {'roles_added': import_result.get('roles_added', 0), 'roles_updated': import_result.get('roles_updated', 0), 'roles_removed': import_result.get('roles_removed', 0), 'relationships_created': import_result.get('relationships_created', 0), 'relationships_removed': import_result.get('relationships_removed', 0), 'tags_assigned': import_result.get('tags_assigned', 0), 'tags_removed': import_result.get('tags_removed', 0), 'errors': import_result.get('errors', []), 'stats': parsed['stats']}})
        except (FileNotFoundError, ValueError) as e:
            logger.warning(f'SIPOC parse error: {e}')
            return jsonify({'success': False, 'error': str(e)})
        except Exception as e:
            logger.exception(f'Error importing SIPOC file: {e}')
            return jsonify({'success': False, 'error': f'Error importing SIPOC file: {str(e)}'})
    finally:
        try:
            if temp_path.exists():
                temp_path.unlink()
        except Exception:
            pass
@roles_bp.route('/api/roles/relationships', methods=['GET'])
@handle_api_errors
def get_role_relationships():
    """
    v4.1.0: Get role relationships (supervisory, tool usage, aliases).

    Query params:
      - role_name: Filter to relationships involving this role
      - type: Filter by relationship type (supervises, uses-tool, alias-of)
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role relationships not available'})
    else:
        role_name = request.args.get('role_name')
        rel_type = request.args.get('type')
        db = get_scan_history_db()
        relationships = db.get_role_relationships(role_name=role_name, rel_type=rel_type)
        return jsonify({'success': True, 'data': {'relationships': relationships, 'total': len(relationships), 'filters': {'role_name': role_name, 'type': rel_type}}})
@roles_bp.route('/api/roles/relationships', methods=['POST'])
@require_csrf
@handle_api_errors
def add_role_relationship():
    """v4.1.0: Add a relationship between two roles."""
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return (jsonify({'success': False, 'error': 'Not available'}), 503)
    else:
        data = request.get_json()
        source = data.get('source_role')
        target = data.get('target_role')
        rel_type = data.get('type', 'inherits-from')
        if not source or not target:
            return (jsonify({'success': False, 'error': 'source_role and target_role required'}), 400)
        else:
            db = get_scan_history_db()
            result = db.add_role_relationship(source, target, rel_type=rel_type, context='manual', import_source='manual')
            return jsonify(result)
@roles_bp.route('/api/roles/relationships/delete', methods=['POST'])
@require_csrf
@handle_api_errors
def delete_role_relationship():
    """v4.1.0: Remove a relationship between two roles."""
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return (jsonify({'success': False, 'error': 'Not available'}), 503)
    else:
        data = request.get_json()
        source = data.get('source_role')
        target = data.get('target_role')
        rel_type = data.get('type')
        if not source or not target:
            return (jsonify({'success': False, 'error': 'source_role and target_role required'}), 400)
        else:
            db = get_scan_history_db()
            result = db.delete_role_relationship(source, target, rel_type=rel_type)
            return jsonify(result)
@roles_bp.route('/api/roles/hierarchy', methods=['GET'])
@handle_api_errors
def get_role_hierarchy():
    """
    v4.1.0: Get full role hierarchy tree for visualization.

    Returns nodes, edges, tree roots, and children map for D3.js rendering.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role hierarchy not available'})
    else:
        db = get_scan_history_db()
        hierarchy = db.get_role_hierarchy()
        return jsonify({'success': True, 'data': hierarchy})
@roles_bp.route('/api/roles/hierarchy/export-html', methods=['GET'])
@handle_api_errors
def export_role_hierarchy_html():
    """
    v4.1.0: Generate interactive HTML export of role hierarchy.

    Query params:
      - org_group: Filter to specific org group(s) (comma-separated)
      - disposition: Filter by disposition (Sanctioned, To Be Retired, TBD)
      - role_type: Filter by role type
      - baselined: \'true\' to include only baselined roles
      - include_tools: \'true\' to include tools (default true)
      - include_all: \'true\' to bypass all filters
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role hierarchy not available'})
    else:
        try:
            from hierarchy_export import generate_hierarchy_html
        except ImportError:
            return jsonify({'success': False, 'error': 'Hierarchy export module not available. Ensure hierarchy_export.py is in the application directory.'})
        db = get_scan_history_db()
        hierarchy = db.get_role_hierarchy()
        relationships = db.get_role_relationships()
        roles = db.get_role_dictionary(include_inactive=True)
        with db.connection() as (_conn, _cursor):
            _cursor.execute('\n        SELECT rft.role_id, rft.function_code, fc.name, fc.color\n        FROM role_function_tags rft\n        LEFT JOIN function_categories fc ON rft.function_code = fc.code\n    ')
            _tag_map = {}
            for row in _cursor.fetchall():
                rid = row[0]
                if rid not in _tag_map:
                    _tag_map[rid] = []
                _tag_map[rid].append({'code': row[1], 'name': row[2] or row[1], 'color': row[3] or '#3b82f6'})
        for role in roles:
            role['function_tags'] = _tag_map.get(role['id'], [])
        filters = {}
        include_all = request.args.get('include_all', 'false').lower() == 'true'
        if not include_all:
            # v4.7.1: Support both org_group (legacy) and functions (new) params
            org_group = request.args.get('org_group', '')
            if org_group:
                filters['org_groups'] = [g.strip() for g in org_group.split(',')]
            functions = request.args.get('functions', '')
            if functions:
                filters['functions'] = [f.strip() for f in functions.split(',')]
            disposition = request.args.get('disposition', '')
            if disposition:
                filters['dispositions'] = [d.strip() for d in disposition.split(',')]
            role_type = request.args.get('role_type', '')
            if role_type:
                filters['role_types'] = [t.strip() for t in role_type.split(',')]
            baselined = request.args.get('baselined', '')
            if baselined.lower() == 'true':
                filters['baselined_only'] = True
            include_tools = request.args.get('include_tools', 'true').lower() == 'true'
            filters['include_tools'] = include_tools
        metadata = {'app_version': '4.1.0', 'export_date': datetime.now().isoformat(), 'filters_applied': filters if not include_all else 'All roles included'}
        try:
            html_content = generate_hierarchy_html(roles=roles, relationships=relationships, hierarchy=hierarchy, filters=filters, metadata=metadata)
            response = make_response(html_content)
            response.headers['Content-Type'] = 'text/html; charset=utf-8'
            # v4.7.2: Allow preview=true to skip download header (for in-browser viewing)
            if request.args.get('preview', '').lower() != 'true':
                response.headers['Content-Disposition'] = f"attachment; filename=AEGIS_Role_Hierarchy_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            return response
        except Exception as e:
            logger.exception(f'Error generating hierarchy HTML: {e}')
            return jsonify({'success': False, 'error': f'Error generating HTML: {str(e)}'})
@roles_bp.route('/api/roles/dictionary/export-template', methods=['GET'])
@handle_api_errors
def export_role_template():
    """
    v4.1.0: Download an interactive HTML template for manual role entry.

    Returns a standalone HTML file with a form, bulk paste, and JSON export.
    Users fill it out offline and import the JSON back into AEGIS.
    """
    try:
        from role_template_export import generate_role_template_html
    except (ImportError, Exception) as e:
        logger.error(f'Failed to import role_template_export: {e}')
        return jsonify({'success': False, 'error': 'Role template module not available. Ensure role_template_export.py is in the application directory.'})
    function_categories = []
    if _shared.SCAN_HISTORY_AVAILABLE:
        try:
            db = get_scan_history_db()
            with db.connection() as (conn, cursor):
                cursor.execute('SELECT code, name, description, color FROM function_categories WHERE is_active = 1 ORDER BY sort_order, name')
                function_categories = [dict(r) for r in cursor.fetchall()]
        except Exception as e:
            logger.warning(f'Error loading function categories: {e}')
    metadata = {'aegis_version': '4.1.0', 'exported_at': datetime.now().isoformat(), 'exported_by': socket.gethostname()}
    try:
        html_content = generate_role_template_html(function_categories=function_categories, metadata=metadata)
        response = make_response(html_content)
        response.headers['Content-Type'] = 'text/html; charset=utf-8'
        response.headers['Content-Disposition'] = f"attachment; filename=AEGIS_Role_Import_Template_{datetime.now().strftime('%Y%m%d')}.html"
        return response
    except Exception as e:
        logger.exception(f'Error generating role template HTML: {e}')
        return jsonify({'success': False, 'error': f'Error generating template: {str(e)}'})
@roles_bp.route('/api/roles/dictionary/clear-sipoc', methods=['POST'])
@require_csrf
@handle_api_errors
def clear_sipoc_import():
    """
    v4.1.0: Clear all SIPOC-imported data for clean reimport.

    Removes roles with source=\'sipoc\', their relationships, and tags.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Role dictionary not available'})
    else:
        db = get_scan_history_db()
        result = db.clear_sipoc_import()
        return jsonify({'success': result.get('success', False), 'data': result})
@roles_bp.route('/api/roles/dictionary/status', methods=['GET'])
@handle_api_errors
def get_dictionary_status():
    """
    Get status of dictionary files and sync state.
    
    Shows:
    - Local database info
    - Master file info (if exists)
    - Shared folder configuration
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Dictionary not available'})
    else:
        db = get_scan_history_db()
        status = db.get_dictionary_status()
        return jsonify({'success': True, 'data': status})
@roles_bp.route('/api/roles/dictionary/export-master', methods=['POST'])
@require_csrf
@handle_api_errors
def export_dictionary_master():
    """
    Export dictionary to a shareable master file.
    
    Creates role_dictionary_master.json that can be distributed to team.
    
    Request body (optional):
    - filepath: Custom output path
    - include_inactive: Include deactivated roles (default false)
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Dictionary not available'})
    else:
        data = request.get_json() or {}
        filepath = data.get('filepath')
        include_inactive = data.get('include_inactive', False)
        db = get_scan_history_db()
        result = db.export_to_master_file(filepath, include_inactive)
        return jsonify(result)
@roles_bp.route('/api/roles/dictionary/create-master', methods=['POST'])
@require_csrf
@handle_api_errors
def create_dictionary_master():
    """
    Create a master dictionary file from current dictionary.
    This is an alias for export-master for clarity.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Dictionary not available'})
    else:
        db = get_scan_history_db()
        result = db.export_to_master_file(include_inactive=False)
        return jsonify(result)
@roles_bp.route('/api/roles/dictionary/sync', methods=['POST'])
@require_csrf
@handle_api_errors
def sync_dictionary():
    """
    Sync dictionary from a master file or scan history.
    
    v2.9.1 D1: Added \'source\' option for sync_from_history
    v2.9.3 B02: Added \'create_if_missing\' option
    
    Request body:
    - source: \'file\' (default) or \'history\'
    - filepath: Path to master file (auto-detected if not provided) - for source=\'file\'
    - merge_mode: \'add_new\' (default), \'replace_all\', or \'update_existing\' - for source=\'file\'
    - create_if_missing: If true and no master file exists, create one from current dictionary
    - min_occurrences: Minimum occurrences for history sync (default: 2)
    - min_confidence: Minimum confidence for history sync (default: 0.7)
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Dictionary not available'})
    else:
        data = request.get_json() or {}
        source = data.get('source', 'file')
        db = get_scan_history_db()
        if source == 'history':
            min_occurrences = data.get('min_occurrences', 2)
            min_confidence = data.get('min_confidence', 0.7)
            result = db.sync_from_history(min_occurrences, min_confidence)
            return jsonify(result)
        else:
            filepath = data.get('filepath')
            merge_mode = data.get('merge_mode', 'add_new')
            create_if_missing = data.get('create_if_missing', False)
            if merge_mode not in ['add_new', 'replace_all', 'update_existing']:
                raise ValidationError(f'Invalid merge_mode: {merge_mode}')
            else:
                result = db.sync_from_master_file(filepath, merge_mode, create_if_missing)
                return jsonify(result)
@roles_bp.route('/api/roles/dictionary/download-master', methods=['GET'])
@handle_api_errors
def download_dictionary_master():
    """
    Download the dictionary as a shareable master JSON file.
    
    This returns the file for download (vs export which saves to server).
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Dictionary not available'})
    else:
        include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
        db = get_scan_history_db()
        roles = db.get_role_dictionary(include_inactive)
        export_data = {'format': 'twr_role_dictionary', 'version': '1.0', 'exported_at': datetime.now().isoformat(), 'exported_by': 'AEGIS', 'role_count': len(roles), 'roles': []}
        for role in roles:
            export_role = {'role_name': role['role_name'], 'aliases': role.get('aliases', []), 'category': role.get('category', 'Role'), 'description': role.get('description'), 'is_deliverable': role.get('is_deliverable', False), 'notes': role.get('notes')}
            export_role = {k: v for k, v in export_role.items() if v is not None}
            export_data['roles'].append(export_role)
        response = make_response(json.dumps(export_data, indent=2))
        response.headers['Content-Type'] = 'application/json'
        response.headers['Content-Disposition'] = 'attachment; filename=role_dictionary_master.json'
        return response
@roles_bp.route('/api/roles/adjudication/export-html', methods=['GET'])
@handle_api_errors
def export_adjudication_html():
    """
    Export adjudication state as a standalone interactive HTML file.

    The HTML file contains an interactive kanban board where users can:
    - Drag and drop roles between status columns
    - Assign function tags
    - Edit categories and notes
    - Generate a JSON import file with decisions
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Adjudication not available'})
    else:
        db = get_scan_history_db()
        all_roles = db.get_all_roles(include_deliverables=True)
        dictionary = db.get_role_dictionary(include_inactive=True)
        with db.connection() as (conn, cursor):
            cursor.execute('SELECT * FROM function_categories WHERE is_active = 1 ORDER BY sort_order, code')
            function_cats = [dict(row) for row in cursor.fetchall()]
        dict_lookup = {}
        for entry in dictionary:
            name = entry.get('role_name', '').lower().strip()
            dict_lookup[name] = entry
        import re as _re
        export_roles = []
        for role in all_roles:
            name = role.get('role_name', '')
            normalized = name.lower().strip()
            dict_entry = dict_lookup.get(normalized, {})
            if dict_entry:
                if not dict_entry.get('is_active', True):
                    status = 'rejected'
                else:
                    if dict_entry.get('is_deliverable', False):
                        status = 'deliverable'
                    else:
                        status = 'confirmed'
            else:
                status = 'pending'
            func_tags = dict_entry.get('function_tags', [])
            tag_codes = []
            if func_tags:
                for t in func_tags:
                    if isinstance(t, dict):
                        tag_codes.append(t.get('code', ''))
                    else:
                        tag_codes.append(str(t))
                tag_codes = [c for c in tag_codes if c]
            conf = 0.5
            doc_count = role.get('unique_document_count', role.get('document_count', 0)) or 0
            mentions = role.get('total_mentions', 0) or 0
            resp_count = role.get('responsibility_count', 0) or 0
            if doc_count >= 3:
                conf += 0.15
            else:
                if doc_count >= 1:
                    conf += 0.05
            if mentions >= 5:
                conf += 0.1
            else:
                if mentions >= 2:
                    conf += 0.05
            if resp_count >= 2:
                conf += 0.15
            else:
                if resp_count >= 1:
                    conf += 0.05
            name_lc = name.lower()
            if _re.search('\\b(engineer|manager|lead|director|officer|specialist|analyst|coordinator|supervisor)\\b', name_lc):
                conf += 0.1
            conf = min(conf, 0.99)
            export_roles.append({'role_name': name, 'status': status, 'category': dict_entry.get('category') or role.get('category', 'Role'), 'confidence': round(conf, 2), 'documents': role.get('documents', []), 'total_mentions': mentions, 'function_tags': tag_codes, 'notes': dict_entry.get('notes', '')})
        categories = []
        for cat in function_cats:
            categories.append({'code': cat.get('code', ''), 'name': cat.get('name', ''), 'description': cat.get('description', ''), 'parent_code': cat.get('parent_code'), 'color': cat.get('color', '#3b82f6'), 'is_active': cat.get('is_active', 1)})
        import socket as _socket
        try:
            from adjudication_export import generate_adjudication_html
        except (ImportError, Exception) as e:
            logger.error(f'Failed to import adjudication_export: {e}')
            return jsonify({'success': False, 'error': 'Adjudication export module not available.'})
        try:
            metadata = {'version': '4.0.3', 'export_date': datetime.now(timezone.utc).isoformat(), 'hostname': _socket.gethostname()}
            html_content = generate_adjudication_html(export_roles, categories, metadata)
            response = make_response(html_content)
            response.headers['Content-Type'] = 'text/html; charset=utf-8'
            date_str = datetime.now().strftime('%Y-%m-%d')
            response.headers['Content-Disposition'] = f'attachment; filename=aegis_adjudication_board_{date_str}.html'
            return response
        except Exception as e:
            logger.error(f'Failed to generate adjudication HTML: {e}')
            return jsonify({'success': False, 'error': f'Failed to generate adjudication export: {str(e)}'})
@roles_bp.route('/api/roles/adjudication/export-pdf', methods=['GET'])
@handle_api_errors
def export_adjudication_pdf():
    """
    Export adjudication state as a formatted PDF report.
    Uses the same data assembly as the HTML export.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Adjudication not available'})
    else:
        db = get_scan_history_db()
        all_roles = db.get_all_roles(include_deliverables=True)
        dictionary = db.get_role_dictionary(include_inactive=True)
        with db.connection() as (conn, cursor):
            cursor.execute('SELECT * FROM function_categories WHERE is_active = 1 ORDER BY sort_order, code')
            function_cats = [dict(row) for row in cursor.fetchall()]
        dict_lookup = {}
        for entry in dictionary:
            name = entry.get('role_name', '').lower().strip()
            dict_lookup[name] = entry
        import re as _re
        export_roles = []
        for role in all_roles:
            name = role.get('role_name', '')
            normalized = name.lower().strip()
            dict_entry = dict_lookup.get(normalized, {})
            if dict_entry:
                if not dict_entry.get('is_active', True):
                    status = 'rejected'
                else:
                    if dict_entry.get('is_deliverable', False):
                        status = 'deliverable'
                    else:
                        status = 'confirmed'
            else:
                status = 'pending'
            func_tags = dict_entry.get('function_tags', [])
            tag_codes = []
            if func_tags:
                for t in func_tags:
                    if isinstance(t, dict):
                        tag_codes.append(t.get('code', ''))
                    else:
                        tag_codes.append(str(t))
                tag_codes = [c for c in tag_codes if c]
            export_roles.append({'role_name': name, 'status': status, 'category': dict_entry.get('category') or role.get('category', 'Role'), 'documents': role.get('documents', []), 'total_mentions': role.get('total_mentions', 0) or 0, 'function_tags': tag_codes, 'notes': dict_entry.get('notes', '')})
        summary = {'pending': 0, 'confirmed': 0, 'deliverable': 0, 'rejected': 0}
        for r in export_roles:
            s = r.get('status', 'pending')
            summary[s] = summary.get(s, 0) + 1
        import socket as _socket
        try:
            from adjudication_report import AdjudicationReportGenerator
        except (ImportError, Exception) as e:
            logger.error(f'Failed to import adjudication_report: {e}')
            return jsonify({'success': False, 'error': 'Adjudication report module not available.'})
        try:
            from config_logging import get_version
            metadata = {'version': get_version(), 'export_date': datetime.now(timezone.utc).isoformat(), 'hostname': _socket.gethostname()}
            generator = AdjudicationReportGenerator()
            pdf_bytes = generator.generate(export_roles, summary, function_cats, metadata)
            response = make_response(pdf_bytes)
            response.headers['Content-Type'] = 'application/pdf'
            date_str = datetime.now().strftime('%Y-%m-%d')
            response.headers['Content-Disposition'] = f'attachment; filename=aegis_adjudication_report_{date_str}.pdf'
            return response
        except Exception as e:
            logger.error(f'Failed to generate adjudication PDF: {e}')
            return jsonify({'success': False, 'error': f'Failed to generate PDF report: {str(e)}'})
@roles_bp.route('/api/roles/adjudication/import-preview', methods=['POST'])
@require_csrf
@handle_api_errors
def preview_import_decisions():
    """
    Preview what an adjudication import would change without applying it.
    Returns a diff showing new roles, changed roles, and unchanged roles.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Adjudication not available'})
    else:
        data = request.get_json()
        if not data:
            raise ValidationError('JSON body required')
        else:
            decisions = data.get('decisions', [])
            if not decisions or not isinstance(decisions, list):
                raise ValidationError('decisions array is required and must not be empty')
            else:
                db = get_scan_history_db()
                dictionary = db.get_role_dictionary(include_inactive=True)
                dict_lookup = {}
                for entry in dictionary:
                    normalized = entry.get('role_name', '').lower().strip()
                    dict_lookup[normalized] = entry
                diff = {'new_roles': [], 'changed': [], 'unchanged': []}
                for decision in decisions:
                    role_name = decision.get('role_name', '').strip()
                    if not role_name:
                        continue
                    else:
                        normalized = role_name.lower().strip()
                        existing = dict_lookup.get(normalized)
                        action = decision.get('action', '')
                        if not existing:
                            diff['new_roles'].append({'role_name': role_name, 'action': action, 'category': decision.get('category', 'Role'), 'function_tags': decision.get('function_tags', [])})
                        else:
                            if not existing.get('is_active', True):
                                current_status = 'rejected'
                            else:
                                if existing.get('is_deliverable', False):
                                    current_status = 'deliverable'
                                else:
                                    current_status = 'confirmed'
                            changes = []
                            if action and action!= current_status:
                                    changes.append(f'status: {current_status} → {action}')
                            if decision.get('category') and decision['category']!= existing.get('category', 'Role'):
                                    changes.append(f"category: {existing.get('category', 'Role')} → {decision['category']}")
                            if decision.get('notes') and decision['notes']!= existing.get('notes', ''):
                                    changes.append('notes updated')
                            if changes:
                                diff['changed'].append({'role_name': role_name, 'changes': changes, 'current_status': current_status, 'new_status': action})
                            else:
                                diff['unchanged'].append({'role_name': role_name, 'status': current_status})
                return jsonify({'success': True, 'diff': diff, 'summary': {'new': len(diff['new_roles']), 'changed': len(diff['changed']), 'unchanged': len(diff['unchanged']), 'total': len(decisions)}})
@roles_bp.route('/api/roles/adjudication/import', methods=['POST'])
@require_csrf
@handle_api_errors
def import_adjudication_decisions():
    """
    Import adjudication decisions from a JSON file (generated by the interactive HTML board).

    Request body:
    {
        \"aegis_version\": \"4.0.3\",
        \"export_type\": \"adjudication_decisions\",
        \"decisions\": [
            {\"role_name\": \"...\", \"action\": \"confirmed|deliverable|rejected|pending\",
             \"category\": \"...\", \"notes\": \"...\", \"function_tags\": [\"CODE1\", \"CODE2\"]}
        ]
    }
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Adjudication not available'})
    else:
        data = request.get_json()
        if not data:
            raise ValidationError('JSON body required')
        else:
            export_type = data.get('export_type')
            if export_type!= 'adjudication_decisions':
                raise ValidationError(f'Invalid export_type: {export_type}. Expected \'adjudication_decisions\'')
            else:
                decisions = data.get('decisions', [])
                if not decisions or not isinstance(decisions, list):
                    raise ValidationError('decisions array is required and must not be empty')
                else:
                    if len(decisions) > 500:
                        raise ValidationError('Maximum 500 decisions per import')
                    else:
                        db = get_scan_history_db()
                        result = db.batch_adjudicate(decisions)
                        return jsonify({'success': True, 'imported': True, 'processed': result.get('processed', 0), 'total': result.get('total', 0), 'errors': result.get('errors', []), 'source': {'aegis_version': data.get('aegis_version'), 'exported_at': data.get('exported_at'), 'exported_by': data.get('exported_by')}})
@roles_bp.route('/api/roles/share/package', methods=['POST'])
@require_csrf
@handle_api_errors
def export_share_package():
    """
    Create a shareable .aegis-roles package file.

    Contains role dictionary with function tags and function categories.
    Returns the package as a downloadable file.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Sharing not available'})
    else:
        db = get_scan_history_db()
        roles = db.get_role_dictionary(include_inactive=False)
        with db.connection() as (conn, cursor):
            cursor.execute('SELECT * FROM function_categories WHERE is_active = 1 ORDER BY sort_order, code')
            function_cats = [dict(row) for row in cursor.fetchall()]
        import socket as _socket
        package_roles = []
        for role in roles:
            r = {'role_name': role['role_name'], 'normalized_name': role.get('normalized_name', role['role_name'].lower().strip()), 'category': role.get('category', 'Role'), 'is_deliverable': role.get('is_deliverable', False), 'notes': role.get('notes', ''), 'aliases': role.get('aliases', [])}
            func_tags = role.get('function_tags', [])
            if func_tags:
                r['function_tags'] = [t.get('code', t) if isinstance(t, dict) else str(t) for t in func_tags]
                r['function_tags'] = [c for c in r['function_tags'] if c]
            package_roles.append(r)
        package_cats = []
        for cat in function_cats:
            package_cats.append({'code': cat.get('code', ''), 'name': cat.get('name', ''), 'description': cat.get('description', ''), 'parent_code': cat.get('parent_code'), 'color': cat.get('color', '#3b82f6'), 'sort_order': cat.get('sort_order', 0)})
        package = {'format': 'aegis_roles_package', 'version': '1.0', 'aegis_version': '4.0.3', 'exported_at': datetime.now(timezone.utc).isoformat(), 'exported_by': _socket.gethostname(), 'role_count': len(package_roles), 'category_count': len(package_cats), 'roles': package_roles, 'function_categories': package_cats}
        response = make_response(json.dumps(package, indent=2))
        response.headers['Content-Type'] = 'application/json'
        date_str = datetime.now().strftime('%Y-%m-%d')
        response.headers['Content-Disposition'] = f'attachment; filename=aegis_roles_{date_str}.aegis-roles'
        return response
@roles_bp.route('/api/roles/share/import-package', methods=['POST'])
@require_csrf
@handle_api_errors
def import_share_package():
    """
    Import a .aegis-roles package from file upload.

    Imports roles (add_new mode) and function categories.
    Accepts multipart/form-data with \'file\' field, or JSON body.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Import not available'})
    else:
        db = get_scan_history_db()
        package_data = None
        if request.content_type and 'multipart/form-data' in request.content_type:
            file = request.files.get('file')
            if not file:
                raise ValidationError('No file uploaded')
            else:
                try:
                    content = file.read().decode('utf-8')
                    package_data = json.loads(content)
                except (UnicodeDecodeError, json.JSONDecodeError) as e:
                    raise ValidationError(f'Invalid package file: {str(e)}')
        else:
            package_data = request.get_json()
        if not package_data:
            raise ValidationError('Package data required')
        else:
            fmt = package_data.get('format')
            if fmt!= 'aegis_roles_package':
                raise ValidationError(f'Invalid package format: {fmt}. Expected \'aegis_roles_package\'')
            else:
                roles = package_data.get('roles', [])
                categories = package_data.get('function_categories', [])
                results = {'roles_added': 0, 'roles_skipped': 0, 'categories_added': 0, 'categories_skipped': 0, 'errors': []}
                if categories:
                    with db.connection() as (conn, cursor):
                        for cat in categories:
                            code = cat.get('code', '').strip()
                            if not code:
                                continue
                            else:
                                try:
                                    cursor.execute('\n                    INSERT INTO function_categories (code, name, description, parent_code, color, sort_order)\n                    VALUES (?, ?, ?, ?, ?, ?)\n                ', (code, cat.get('name', code), cat.get('description', ''), cat.get('parent_code'), cat.get('color', '#3b82f6'), cat.get('sort_order', 0)))
                                    results['categories_added'] += 1
                                except sqlite3.IntegrityError:
                                    results['categories_skipped'] += 1
                                except Exception as e:
                                    results['errors'].append(f'Category {code}: {str(e)}')
                if roles:
                    import_result = db.import_roles_to_dictionary(roles, source='package_import', source_document=f"package_{package_data.get('exported_by', 'unknown')}", created_by='package_import')
                    results['roles_added'] = import_result.get('added', 0)
                    results['roles_skipped'] = import_result.get('skipped', 0)
                    if import_result.get('errors'):
                        results['errors'].extend(import_result['errors'])
                results['success'] = True
                results['source'] = {'aegis_version': package_data.get('aegis_version'), 'exported_at': package_data.get('exported_at'), 'exported_by': package_data.get('exported_by')}
                return jsonify(results)
@roles_bp.route('/api/sow/data', methods=['GET'])
@handle_api_errors
def get_sow_data():
    """
    Fetch all available data for SOW configuration.
    Returns roles, statements, documents, function categories, and relationships.
    """
    if not _shared.SCAN_HISTORY_AVAILABLE:
        return jsonify({'success': False, 'error': 'Scan history not available'})
    else:
        from scan_history import get_scan_history_db
        db = get_scan_history_db()
        dictionary = db.get_role_dictionary(include_inactive=False)
        with db.connection() as (conn, cursor):
            cursor.execute('SELECT * FROM function_categories WHERE is_active = 1 ORDER BY sort_order, code')
            function_cats = [dict(row) for row in cursor.fetchall()]
            cursor.execute('\n        SELECT rft.role_id, rft.role_name, rft.function_code,\n               fc.name as function_name, fc.color as function_color\n        FROM role_function_tags rft\n        LEFT JOIN function_categories fc ON fc.code = rft.function_code\n    ')
            tag_rows = cursor.fetchall()
            tag_lookup = {}
            for tr in tag_rows:
                rn = tr['role_name'].lower().strip()
                if rn not in tag_lookup:
                    tag_lookup[rn] = []
                tag_lookup[rn].append({'code': tr['function_code'], 'name': tr['function_name'] or tr['function_code'], 'color': tr['function_color'] or '#3b82f6'})
            enriched_roles = []
            for role in dictionary:
                rn = role.get('role_name', '').lower().strip()
                role['function_tags'] = tag_lookup.get(rn, [])
                enriched_roles.append(role)
            cursor.execute('SELECT * FROM role_relationships')
            relationships = [dict(row) for row in cursor.fetchall()]
            cursor.execute('\n        SELECT d.id, d.filename, d.filepath, d.word_count, d.paragraph_count,\n               COUNT(s.id) as scan_count,\n               MAX(s.score) as latest_score,\n               MAX(s.grade) as latest_grade\n        FROM documents d\n        LEFT JOIN scans s ON s.document_id = d.id\n        GROUP BY d.id\n        ORDER BY d.filename\n    ')
            documents = [dict(row) for row in cursor.fetchall()]
            cursor.execute('\n        SELECT ss.*, d.filename as source_document\n        FROM scan_statements ss\n        JOIN scans s ON s.id = ss.scan_id\n        JOIN documents d ON d.id = s.document_id\n        WHERE s.id IN (\n            SELECT MAX(id) FROM scans GROUP BY document_id\n        )\n        ORDER BY ss.document_id, ss.position_index\n    ')
            statements = [dict(row) for row in cursor.fetchall()]
        return jsonify({'success': True, 'data': {'roles': enriched_roles, 'statements': statements, 'documents': documents, 'function_categories': function_cats, 'relationships': relationships, 'counts': {'roles': len(enriched_roles), 'statements': len(statements), 'documents': len(documents), 'categories': len(function_cats), 'relationships': len(relationships)}}})